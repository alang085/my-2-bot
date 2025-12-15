"""测试所有Excel报表功能"""
import asyncio
import sys
import os
from pathlib import Path
from datetime import datetime, timedelta
import pytz

# 添加项目根目录到路径
project_root = Path(__file__).parent.absolute()
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

# 导入模块
import init_db
import db_operations
from utils.incremental_report_generator import get_or_create_baseline_date, prepare_incremental_data
from utils.excel_export import (
    export_incremental_orders_report_to_excel,
    export_daily_changes_to_excel
)
from utils.daily_report_generator import calculate_daily_summary

BEIJING_TZ = pytz.timezone('Asia/Shanghai')


async def test_incremental_excel():
    """测试增量订单报表Excel"""
    print("=" * 60)
    print("测试1: 增量订单报表Excel")
    print("=" * 60)
    
    try:
        # 获取基准日期
        baseline_date = await get_or_create_baseline_date()
        current_date = datetime.now(BEIJING_TZ).strftime('%Y-%m-%d')
        
        print(f"基准日期: {baseline_date}")
        print(f"当前日期: {current_date}")
        
        # 准备增量数据
        incremental_data = await prepare_incremental_data(baseline_date)
        orders_data = incremental_data.get('orders', [])
        expense_records = incremental_data.get('expenses', [])
        
        print(f"增量订单数: {len(orders_data)}")
        print(f"增量开销记录数: {len(expense_records)}")
        
        # 生成Excel报表
        excel_path = await export_incremental_orders_report_to_excel(
            baseline_date,
            current_date,
            orders_data,
            expense_records
        )
        
        if os.path.exists(excel_path):
            file_size = os.path.getsize(excel_path) / 1024
            print(f"✅ Excel报表已生成: {excel_path}")
            print(f"✅ 文件大小: {file_size:.2f} KB")
            return excel_path
        else:
            print(f"❌ Excel文件不存在: {excel_path}")
            return None
            
    except Exception as e:
        print(f"❌ 测试失败: {e}")
        import traceback
        traceback.print_exc()
        return None


async def test_daily_changes_excel():
    """测试每日数据变更Excel"""
    print("\n" + "=" * 60)
    print("测试2: 每日数据变更Excel")
    print("=" * 60)
    
    try:
        # 使用当前日期
        current_date = datetime.now(BEIJING_TZ).strftime('%Y-%m-%d')
        print(f"查询日期: {current_date}")
        
        # 生成Excel报表
        excel_path = await export_daily_changes_to_excel(current_date)
        
        if os.path.exists(excel_path):
            file_size = os.path.getsize(excel_path) / 1024
            print(f"✅ Excel报表已生成: {excel_path}")
            print(f"✅ 文件大小: {file_size:.2f} KB")
            return excel_path
        else:
            print(f"❌ Excel文件不存在: {excel_path}")
            return None
            
    except Exception as e:
        print(f"❌ 测试失败: {e}")
        import traceback
        traceback.print_exc()
        return None


async def test_multiple_dates_excel():
    """测试多日期Excel报表"""
    print("\n" + "=" * 60)
    print("测试3: 多日期Excel报表")
    print("=" * 60)
    
    try:
        # 测试最近3天的报表
        dates = []
        for i in range(3):
            date = (datetime.now(BEIJING_TZ) - timedelta(days=i)).strftime('%Y-%m-%d')
            dates.append(date)
        
        print(f"测试日期: {', '.join(dates)}")
        
        excel_paths = []
        for date in dates:
            try:
                excel_path = await export_daily_changes_to_excel(date)
                if os.path.exists(excel_path):
                    file_size = os.path.getsize(excel_path) / 1024
                    print(f"✅ {date}: {file_size:.2f} KB")
                    excel_paths.append(excel_path)
                else:
                    print(f"⚠️  {date}: 文件不存在")
            except Exception as e:
                print(f"❌ {date}: {str(e)}")
        
        print(f"\n✅ 成功生成 {len(excel_paths)} 个Excel文件")
        return excel_paths
        
    except Exception as e:
        print(f"❌ 测试失败: {e}")
        import traceback
        traceback.print_exc()
        return []


async def verify_excel_structure(excel_path):
    """验证Excel文件结构"""
    print("\n" + "=" * 60)
    print("验证Excel文件结构")
    print("=" * 60)
    
    try:
        from openpyxl import load_workbook
        
        if not os.path.exists(excel_path):
            print(f"❌ 文件不存在: {excel_path}")
            return False
        
        wb = load_workbook(excel_path)
        
        print(f"✅ Excel文件已打开")
        print(f"工作表数量: {len(wb.sheetnames)}")
        print(f"工作表列表: {', '.join(wb.sheetnames)}")
        
        # 检查每个工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n📊 工作表: {sheet_name}")
            print(f"  行数: {ws.max_row}")
            print(f"  列数: {ws.max_column}")
            
            # 显示表头（第一行）
            if ws.max_row > 0:
                headers = []
                for col in range(1, min(ws.max_column + 1, 10)):
                    cell_value = ws.cell(row=1, column=col).value
                    if cell_value:
                        headers.append(str(cell_value))
                print(f"  表头: {', '.join(headers)}")
        
        wb.close()
        print("\n✅ Excel文件结构验证通过")
        return True
        
    except Exception as e:
        print(f"❌ 验证失败: {e}")
        import traceback
        traceback.print_exc()
        return False


async def list_excel_files():
    """列出所有生成的Excel文件"""
    print("\n" + "=" * 60)
    print("生成的Excel文件列表")
    print("=" * 60)
    
    temp_dir = os.path.join(project_root, 'temp')
    if not os.path.exists(temp_dir):
        print(f"⚠️  temp目录不存在: {temp_dir}")
        return
    
    excel_files = []
    for file in os.listdir(temp_dir):
        if file.endswith('.xlsx'):
            file_path = os.path.join(temp_dir, file)
            file_size = os.path.getsize(file_path) / 1024
            mtime = datetime.fromtimestamp(os.path.getmtime(file_path))
            excel_files.append({
                'name': file,
                'path': file_path,
                'size': file_size,
                'mtime': mtime
            })
    
    if excel_files:
        # 按修改时间排序
        excel_files.sort(key=lambda x: x['mtime'], reverse=True)
        
        print(f"找到 {len(excel_files)} 个Excel文件:\n")
        for i, file_info in enumerate(excel_files, 1):
            print(f"{i}. {file_info['name']}")
            print(f"   大小: {file_info['size']:.2f} KB")
            print(f"   修改时间: {file_info['mtime'].strftime('%Y-%m-%d %H:%M:%S')}")
            print(f"   路径: {file_info['path']}\n")
    else:
        print("⚠️  未找到Excel文件")


async def main():
    """主测试函数"""
    print("\n" + "=" * 60)
    print("Excel报表功能完整测试")
    print("=" * 60 + "\n")
    
    try:
        # 初始化数据库
        print("初始化数据库...")
        init_db.init_database()
        print("✅ 数据库初始化完成\n")
        
        # 测试1: 增量订单报表Excel
        incremental_excel = await test_incremental_excel()
        if incremental_excel:
            await verify_excel_structure(incremental_excel)
        
        # 测试2: 每日数据变更Excel
        daily_excel = await test_daily_changes_excel()
        if daily_excel:
            await verify_excel_structure(daily_excel)
        
        # 测试3: 多日期Excel报表
        multiple_excels = await test_multiple_dates_excel()
        
        # 列出所有Excel文件
        await list_excel_files()
        
        print("\n" + "=" * 60)
        print("✅ 所有测试完成！")
        print("=" * 60)
        print("\n💡 提示：")
        print("1. 打开temp目录查看生成的Excel文件")
        print("2. 检查Excel文件格式和内容")
        print("3. 测试利息明细的展开/折叠功能")
        print("4. 验证数据准确性")
        
    except Exception as e:
        print(f"\n❌ 测试失败: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    asyncio.run(main())

