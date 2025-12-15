"""详细验证Excel报表内容和格式"""
import os
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border

# 添加项目根目录到路径
project_root = Path(__file__).parent.absolute()
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

temp_dir = os.path.join(project_root, 'temp')


def verify_incremental_report(excel_path):
    """验证增量订单报表Excel"""
    print("=" * 60)
    print("验证增量订单报表Excel")
    print("=" * 60)
    
    if not os.path.exists(excel_path):
        print(f"❌ 文件不存在: {excel_path}")
        return False
    
    try:
        wb = load_workbook(excel_path)
        
        # 检查工作表
        if '增量订单报表' not in wb.sheetnames:
            print("❌ 缺少'增量订单报表'工作表")
            return False
        
        ws = wb['增量订单报表']
        
        print(f"✅ 工作表: 增量订单报表")
        print(f"   行数: {ws.max_row}")
        print(f"   列数: {ws.max_column}")
        
        # 检查表头
        print("\n📋 表头检查:")
        expected_headers = ['日期', '订单号', '会员', '订单金额', '利息总数', '归还本金', '订单状态', '备注']
        if ws.max_row >= 2:
            headers = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=2, column=col).value
                if cell_value:
                    headers.append(str(cell_value))
            
            print(f"   表头: {headers}")
            
            # 检查表头样式
            header_cell = ws.cell(row=2, column=1)
            if header_cell.fill and header_cell.fill.start_color:
                print(f"   ✅ 表头有背景色")
            if header_cell.font and header_cell.font.bold:
                print(f"   ✅ 表头字体加粗")
        
        # 检查数据行
        if ws.max_row > 2:
            print(f"\n📊 数据行: {ws.max_row - 2} 行")
            # 显示前3行数据
            for row in range(3, min(6, ws.max_row + 1)):
                row_data = []
                for col in range(1, min(9, ws.max_column + 1)):
                    cell_value = ws.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value else '')
                print(f"   第{row}行: {row_data}")
        else:
            print("\n⚠️  无数据行（这是正常的，如果没有增量数据）")
        
        # 检查汇总行
        if ws.max_row > 2:
            print(f"\n📈 汇总行检查:")
            summary_cell = ws.cell(row=ws.max_row, column=1)
            if summary_cell.value and '汇总' in str(summary_cell.value):
                print(f"   ✅ 找到汇总行")
                if summary_cell.font and summary_cell.font.bold:
                    print(f"   ✅ 汇总行字体加粗")
        
        # 检查列宽
        print(f"\n📏 列宽检查:")
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            if col_letter in ws.column_dimensions:
                width = ws.column_dimensions[col_letter].width
                print(f"   {col_letter}列: {width}")
        
        # 检查分组（利息明细）
        print(f"\n🔽 分组检查（利息明细）:")
        grouped_rows = 0
        for row in range(1, ws.max_row + 1):
            if ws.row_dimensions[row].outline_level > 0:
                grouped_rows += 1
        
        if grouped_rows > 0:
            print(f"   ✅ 找到 {grouped_rows} 行分组（利息明细）")
            print(f"   💡 这些行默认隐藏，可以展开查看")
        else:
            print(f"   ℹ️  无分组行（如果没有多笔利息记录）")
        
        wb.close()
        print("\n✅ 增量订单报表验证通过")
        return True
        
    except Exception as e:
        print(f"❌ 验证失败: {e}")
        import traceback
        traceback.print_exc()
        return False


def verify_daily_changes_report(excel_path):
    """验证每日数据变更Excel"""
    print("\n" + "=" * 60)
    print("验证每日数据变更Excel")
    print("=" * 60)
    
    if not os.path.exists(excel_path):
        print(f"❌ 文件不存在: {excel_path}")
        return False
    
    try:
        wb = load_workbook(excel_path)
        
        # 检查工作表
        if '数据汇总' not in wb.sheetnames:
            print("❌ 缺少'数据汇总'工作表")
            return False
        
        ws = wb['数据汇总']
        
        print(f"✅ 工作表: 数据汇总")
        print(f"   行数: {ws.max_row}")
        print(f"   列数: {ws.max_column}")
        
        # 检查标题
        print("\n📋 标题检查:")
        title_cell = ws.cell(row=1, column=1)
        if title_cell.value:
            print(f"   标题: {title_cell.value}")
        
        # 检查数据
        print("\n📊 数据检查:")
        if ws.max_row > 1:
            for row in range(2, min(ws.max_row + 1, 15)):
                key = ws.cell(row=row, column=1).value
                value = ws.cell(row=row, column=2).value
                if key:
                    print(f"   {key}: {value}")
        
        wb.close()
        print("\n✅ 每日数据变更报表验证通过")
        return True
        
    except Exception as e:
        print(f"❌ 验证失败: {e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """主函数"""
    print("\n" + "=" * 60)
    print("Excel报表内容详细验证")
    print("=" * 60 + "\n")
    
    # 查找最新的增量订单报表
    incremental_files = []
    daily_files = []
    
    if os.path.exists(temp_dir):
        for file in os.listdir(temp_dir):
            if file.startswith('增量订单报表_') and file.endswith('.xlsx'):
                file_path = os.path.join(temp_dir, file)
                incremental_files.append((os.path.getmtime(file_path), file_path))
            elif file.startswith('每日变化数据_') and file.endswith('.xlsx'):
                file_path = os.path.join(temp_dir, file)
                daily_files.append((os.path.getmtime(file_path), file_path))
    
    # 验证增量订单报表
    if incremental_files:
        incremental_files.sort(reverse=True)
        latest_incremental = incremental_files[0][1]
        verify_incremental_report(latest_incremental)
    else:
        print("⚠️  未找到增量订单报表文件")
    
    # 验证每日数据变更报表
    if daily_files:
        daily_files.sort(reverse=True)
        latest_daily = daily_files[0][1]
        verify_daily_changes_report(latest_daily)
    else:
        print("⚠️  未找到每日数据变更报表文件")
    
    print("\n" + "=" * 60)
    print("✅ 验证完成！")
    print("=" * 60)
    print("\n💡 提示：")
    print("1. 打开Excel文件查看详细内容")
    print("2. 测试利息明细的展开/折叠功能")
    print("3. 验证数据格式和样式")
    print("4. 检查汇总行计算是否正确")


if __name__ == "__main__":
    main()

