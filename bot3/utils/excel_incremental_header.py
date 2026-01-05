"""增量订单报表 - 标题和表头模块

包含创建Excel标题和表头的逻辑。
"""

from openpyxl import Workbook
from openpyxl.styles import Font


def create_sheet_header(
    ws_orders, baseline_date: str, current_date: str, styles: dict
) -> None:
    """创建工作表标题和表头

    Args:
        ws_orders: Excel工作表对象
        baseline_date: 基准日期
        current_date: 当前日期
        styles: 样式字典
    """
    # 标题
    ws_orders.merge_cells("A1:H1")
    title_text = f"增量订单报表 (基准日期: {baseline_date}, 当前日期: {current_date})"
    ws_orders["A1"] = title_text
    ws_orders["A1"].font = styles["title_font"]
    ws_orders["A1"].alignment = styles["center_align"]

    # 表头
    headers = [
        "日期",
        "订单号",
        "会员",
        "订单金额",
        "利息总数",
        "归还本金",
        "订单状态",
        "备注",
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_orders.cell(row=2, column=col_idx, value=header)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center_align"]
        cell.border = styles["border"]
