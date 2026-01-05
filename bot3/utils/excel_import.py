"""Excel导入工具 - 从Excel文件导入订单数据"""

import logging
import os
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook

from utils.excel_import_columns import parse_column_indices
from utils.excel_import_row_parse import parse_order_row
from utils.excel_import_worksheet import find_header_row, find_worksheet

logger = logging.getLogger(__name__)


def _load_and_validate_worksheet(excel_file_path: str) -> Tuple[Any, bool]:
    """加载并验证工作表

    Args:
        excel_file_path: Excel文件路径

    Returns:
        (工作表对象, 是否找到)
    """
    if not os.path.exists(excel_file_path):
        logger.error(f"Excel文件不存在: {excel_file_path}")
        return None, False

    try:
        wb = load_workbook(excel_file_path, data_only=True)
        ws, found = find_worksheet(wb)
        return ws, found
    except Exception as e:
        logger.error(f"加载Excel文件失败: {e}", exc_info=True)
        return None, False


def _parse_orders_from_worksheet(
    ws: Any, header_row: int, col_indices: Dict[str, int]
) -> List[Dict[str, Any]]:
    """从工作表解析订单数据

    Args:
        ws: 工作表对象
        header_row: 表头行号
        col_indices: 列索引字典

    Returns:
        订单数据列表
    """
    orders = []
    data_start_row = header_row + 1

    for row_idx in range(data_start_row, ws.max_row + 1):
        row = ws[row_idx]
        row_values = [cell.value for cell in row]

        if all(v is None or str(v).strip() == "" for v in row_values):
            continue

        order_data = parse_order_row(row_values, col_indices)
        if order_data:
            orders.append(order_data)

    return orders


def parse_excel_orders(excel_file_path: str) -> List[Dict[str, Any]]:
    """
    从Excel文件解析订单数据

    Args:
        excel_file_path: Excel文件路径

    Returns:
        List[Dict]: 订单数据列表，每个订单包含：
            - order_id: 订单号
            - date: 日期
            - customer: 会员名
            - group_id: 归属ID
            - amount: 订单金额
            - state: 订单状态
            - weekday_group: 星期分组（从日期计算）
    """
    ws, found = _load_and_validate_worksheet(excel_file_path)
    if not found:
        return []

    header_row = find_header_row(ws)
    if header_row is None:
        logger.error("未找到表头行")
        return []

    headers = [str(cell.value or "").strip() for cell in ws[header_row]]
    col_indices = parse_column_indices(headers)

    orders = _parse_orders_from_worksheet(ws, header_row, col_indices)
    logger.info(f"从Excel文件解析到 {len(orders)} 条订单")
    return orders


def _initialize_import_result() -> Dict[str, Any]:
    """初始化导入结果

    Returns:
        结果字典
    """
    return {
        "success": True,
        "total": 0,
        "imported": 0,
        "failed": 0,
        "errors": [],
    }


async def _import_single_order(
    order_data: Dict[str, Any], chat_id: int, result: Dict[str, Any]
) -> None:
    """导入单个订单

    Args:
        order_data: 订单数据
        chat_id: 分配的chat_id
        result: 结果字典（会被修改）
    """
    import db_operations

    try:
        existing_order = await db_operations.get_order_by_order_id(
            order_data["order_id"]
        )
        if existing_order:
            logger.warning(f"订单已存在，跳过: {order_data['order_id']}")
            result["failed"] += 1
            result["errors"].append(f"订单 {order_data['order_id']} 已存在")
            return

        order_data["chat_id"] = chat_id
        order_data["group"] = order_data.pop("weekday_group")

        success = await db_operations.create_order(order_data)
        if success:
            result["imported"] += 1
            logger.info(f"成功导入订单: {order_data['order_id']}")
        else:
            result["failed"] += 1
            result["errors"].append(f"导入订单失败: {order_data['order_id']}")

    except Exception as e:
        result["failed"] += 1
        error_msg = f"导入订单 {order_data.get('order_id', '未知')} 时出错: {e}"
        result["errors"].append(error_msg)
        logger.error(error_msg, exc_info=True)


async def import_orders_from_excel(
    excel_file_path: str, chat_id_base: int = -1000000000000
) -> Dict[str, Any]:
    """
    从Excel文件导入订单到数据库

    Args:
        excel_file_path: Excel文件路径
        chat_id_base: chat_id的基础值（每个订单会递增）

    Returns:
        Dict: 导入结果
            - success: bool
            - total: int - 总订单数
            - imported: int - 成功导入数
            - failed: int - 失败数
            - errors: List[str] - 错误信息列表
    """
    result = _initialize_import_result()
    orders = parse_excel_orders(excel_file_path)
    result["total"] = len(orders)

    if not orders:
        result["success"] = False
        result["errors"].append("Excel文件中没有找到订单数据")
        return result

    chat_id_counter = chat_id_base
    for order_data in orders:
        chat_id_counter -= 1
        await _import_single_order(order_data, chat_id_counter, result)

    result["success"] = result["failed"] == 0
    return result
