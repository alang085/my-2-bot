"""订单工作表创建模块（订单总表Excel）

包含订单总表Excel中各个工作表的创建函数。
"""

# 标准库
from typing import Dict, List

# 第三方库
from openpyxl import Workbook
from openpyxl.styles import Font

# 本地模块
from constants import ORDER_STATES
from utils.excel_export import format_datetime_to_beijing


def create_orders_sheet(wb: Workbook, orders: List[Dict], styles: Dict) -> None:
    """创建订单总表工作表"""
    from utils.excel_orders_data import write_orders_data
    from utils.excel_orders_header import create_orders_header
    from utils.excel_orders_summary import (adjust_orders_column_width,
                                            create_orders_summary)

    ws_orders = wb.create_sheet("订单总表", 0)

    # 创建表头
    create_orders_header(ws_orders, styles)

    # 写入数据行
    row_idx, total_amount, total_interest_all = write_orders_data(
        ws_orders, orders, styles
    )

    # 创建汇总行
    create_orders_summary(
        ws_orders, orders, row_idx, total_amount, total_interest_all, styles
    )

    # 调整列宽
    adjust_orders_column_width(ws_orders)


def create_completed_orders_sheet(
    wb: Workbook, completed_orders: List[Dict], styles: Dict
) -> None:
    """创建完成订单工作表（显示所有完成订单的总计）"""
    if not completed_orders:
        return

    ws_completed = wb.create_sheet("完成订单")
    ws_completed.merge_cells("A1:D1")
    ws_completed["A1"] = "完成订单（总计）"
    ws_completed["A1"].font = styles["title_font"]
    ws_completed["A1"].alignment = styles["center_align"]

    headers = ["时间", "订单号", "金额", "完成时间"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_completed.cell(row=2, column=col_idx, value=header)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center_align"]
        cell.border = styles["border"]

    row_idx = 3
    total_amount = 0.0
    for order in completed_orders:
        date_str = order.get("date", "")[:10] if order.get("date") else "未知"
        order_id = order.get("order_id", "未知")
        amount = float(order.get("amount", 0) or 0)
        total_amount += amount
        updated_at_raw = (
            order.get("updated_at", "") if order.get("updated_at") else "未知"
        )
        updated_at = (
            format_datetime_to_beijing(updated_at_raw)
            if updated_at_raw != "未知"
            else "未知"
        )

        ws_completed.cell(row=row_idx, column=1, value=date_str).border = styles[
            "border"
        ]
        ws_completed.cell(row=row_idx, column=2, value=order_id).border = styles[
            "border"
        ]
        amount_cell = ws_completed.cell(row=row_idx, column=3, value=amount)
        amount_cell.border = styles["border"]
        amount_cell.number_format = "#,##0.00"
        amount_cell.alignment = styles["right_align"]
        ws_completed.cell(row=row_idx, column=4, value=updated_at).border = styles[
            "border"
        ]

        row_idx += 1

    # 添加汇总行
    from openpyxl.styles import Font

    summary_row = row_idx + 1
    ws_completed.cell(row=summary_row, column=1, value="汇总").font = Font(bold=True)
    ws_completed.cell(row=summary_row, column=1).border = styles["border"]
    ws_completed.cell(
        row=summary_row, column=2, value=f"订单数: {len(completed_orders)}"
    ).font = Font(bold=True)
    ws_completed.cell(row=summary_row, column=2).border = styles["border"]
    total_cell = ws_completed.cell(row=summary_row, column=3, value=total_amount)
    total_cell.font = Font(bold=True)
    total_cell.border = styles["border"]
    total_cell.number_format = "#,##0.00"
    total_cell.alignment = styles["right_align"]
    ws_completed.cell(row=summary_row, column=4, value="").border = styles["border"]

    ws_completed.column_dimensions["A"].width = 12
    ws_completed.column_dimensions["B"].width = 15
    ws_completed.column_dimensions["C"].width = 15
    ws_completed.column_dimensions["D"].width = 20


def create_breach_orders_sheet(
    wb: Workbook, breach_orders: List[Dict], styles: Dict
) -> None:
    """创建违约订单工作表"""
    if not breach_orders:
        return

    ws_breach_active = wb.create_sheet("违约订单")
    ws_breach_active.merge_cells("A1:D1")
    ws_breach_active["A1"] = "违约订单（当日有变动）"
    ws_breach_active["A1"].font = styles["title_font"]
    ws_breach_active["A1"].alignment = styles["center_align"]

    headers = ["时间", "订单号", "金额", "违约时间"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_breach_active.cell(row=2, column=col_idx, value=header)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center_align"]
        cell.border = styles["border"]

    row_idx = 3
    for order in breach_orders:
        date_str = order.get("date", "")[:10] if order.get("date") else "未知"
        order_id = order.get("order_id", "未知")
        amount = order.get("amount", 0)
        updated_at_raw = (
            order.get("updated_at", "") if order.get("updated_at") else "未知"
        )
        updated_at = (
            format_datetime_to_beijing(updated_at_raw)
            if updated_at_raw != "未知"
            else "未知"
        )

        ws_breach_active.cell(row=row_idx, column=1, value=date_str).border = styles[
            "border"
        ]
        ws_breach_active.cell(row=row_idx, column=2, value=order_id).border = styles[
            "border"
        ]
        amount_cell = ws_breach_active.cell(
            row=row_idx, column=3, value=float(amount) if amount else 0
        )
        amount_cell.border = styles["border"]
        amount_cell.number_format = "#,##0.00"
        amount_cell.alignment = styles["right_align"]
        ws_breach_active.cell(row=row_idx, column=4, value=updated_at).border = styles[
            "border"
        ]

        row_idx += 1

    ws_breach_active.column_dimensions["A"].width = 12
    ws_breach_active.column_dimensions["B"].width = 15
    ws_breach_active.column_dimensions["C"].width = 15
    ws_breach_active.column_dimensions["D"].width = 20


def create_breach_end_orders_sheet(
    wb: Workbook, breach_end_orders: List[Dict], styles: Dict
) -> None:
    """创建违约完成订单工作表"""
    if not breach_end_orders:
        return

    ws_breach = wb.create_sheet("违约完成订单")
    ws_breach.merge_cells("A1:D1")
    ws_breach["A1"] = "违约完成订单（当日有变动）"
    ws_breach["A1"].font = styles["title_font"]
    ws_breach["A1"].alignment = styles["center_align"]

    headers = ["时间", "订单号", "金额", "完成时间"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_breach.cell(row=2, column=col_idx, value=header)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center_align"]
        cell.border = styles["border"]

    row_idx = 3
    for order in breach_end_orders:
        date_str = order.get("date", "")[:10] if order.get("date") else "未知"
        order_id = order.get("order_id", "未知")
        amount = order.get("amount", 0)
        updated_at_raw = (
            order.get("updated_at", "") if order.get("updated_at") else "未知"
        )
        updated_at = (
            format_datetime_to_beijing(updated_at_raw)
            if updated_at_raw != "未知"
            else "未知"
        )

        ws_breach.cell(row=row_idx, column=1, value=date_str).border = styles["border"]
        ws_breach.cell(row=row_idx, column=2, value=order_id).border = styles["border"]
        amount_cell = ws_breach.cell(
            row=row_idx, column=3, value=float(amount) if amount else 0
        )
        amount_cell.border = styles["border"]
        amount_cell.number_format = "#,##0.00"
        amount_cell.alignment = styles["right_align"]
        ws_breach.cell(row=row_idx, column=4, value=updated_at).border = styles[
            "border"
        ]

        row_idx += 1

    ws_breach.column_dimensions["A"].width = 12
    ws_breach.column_dimensions["B"].width = 15
    ws_breach.column_dimensions["C"].width = 15
    ws_breach.column_dimensions["D"].width = 20


def create_daily_summary_sheet(wb: Workbook, daily_summary: Dict, styles: Dict) -> None:
    """创建日切数据汇总工作表"""
    if not daily_summary:
        return

    ws_summary = wb.create_sheet("日切数据汇总")
    ws_summary.merge_cells("A1:B1")
    ws_summary["A1"] = "日切数据汇总"
    ws_summary["A1"].font = styles["title_font"]
    ws_summary["A1"].alignment = styles["center_align"]

    summary_data = [
        ["新增订单数", daily_summary.get("new_orders_count", 0)],
        ["新增订单金额", daily_summary.get("new_orders_amount", 0.0)],
        ["完结订单数", daily_summary.get("completed_orders_count", 0)],
        ["完结订单金额", daily_summary.get("completed_orders_amount", 0.0)],
        ["违约完成数", daily_summary.get("breach_end_orders_count", 0)],
        ["违约完成金额", daily_summary.get("breach_end_orders_amount", 0.0)],
        ["当日利息", daily_summary.get("daily_interest", 0.0)],
        ["公司开销", daily_summary.get("company_expenses", 0.0)],
        ["其他开销", daily_summary.get("other_expenses", 0.0)],
        [
            "总开销",
            daily_summary.get("company_expenses", 0.0)
            + daily_summary.get("other_expenses", 0.0),
        ],
    ]

    row_idx = 3
    for label, value in summary_data:
        label_cell = ws_summary.cell(row=row_idx, column=1, value=label)
        label_cell.font = Font(bold=True)
        label_cell.border = styles["border"]
        value_cell = ws_summary.cell(row=row_idx, column=2, value=value)
        if isinstance(value, float):
            value_cell.number_format = "#,##0.00"
            value_cell.alignment = styles["right_align"]
        else:
            value_cell.alignment = styles["center_align"]
        value_cell.border = styles["border"]
        row_idx += 1

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 20


def create_chat_id_mapping_sheet(
    wb: Workbook, all_orders_for_chat_id: List[Dict], styles: Dict
) -> None:
    """创建订单chat_id对应表工作表"""
    if not all_orders_for_chat_id:
        return

    ws_chat_ids = wb.create_sheet("订单chat_id对应表")
    ws_chat_ids.merge_cells("A1:C1")
    ws_chat_ids["A1"] = "订单chat_id对应表（所有订单）"
    ws_chat_ids["A1"].font = styles["title_font"]
    ws_chat_ids["A1"].alignment = styles["center_align"]

    headers = ["订单号", "chat_id", "状态"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_chat_ids.cell(row=2, column=col_idx, value=header)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center_align"]
        cell.border = styles["border"]

    row_idx = 3
    for order in all_orders_for_chat_id:
        order_id = order.get("order_id", "未知")
        chat_id = order.get("chat_id", "")
        state = ORDER_STATES.get(order.get("state", ""), order.get("state", "未知"))

        ws_chat_ids.cell(row=row_idx, column=1, value=order_id).border = styles[
            "border"
        ]
        chat_id_value = chat_id if chat_id else "无"
        ws_chat_ids.cell(row=row_idx, column=2, value=chat_id_value).border = styles[
            "border"
        ]
        ws_chat_ids.cell(row=row_idx, column=3, value=state).border = styles["border"]
        row_idx += 1

    # 汇总行
    if all_orders_for_chat_id:
        ws_chat_ids.merge_cells(f"A{row_idx}:B{row_idx}")
        summary_cell = ws_chat_ids.cell(
            row=row_idx, column=1, value=f"总计: {len(all_orders_for_chat_id)} 个订单"
        )
        summary_cell.font = Font(bold=True)
        summary_cell.alignment = styles["right_align"]
        summary_cell.border = styles["border"]
        ws_chat_ids.cell(row=row_idx, column=3, value="").border = styles["border"]

    ws_chat_ids.column_dimensions["A"].width = 18
    ws_chat_ids.column_dimensions["B"].width = 20
    ws_chat_ids.column_dimensions["C"].width = 12
