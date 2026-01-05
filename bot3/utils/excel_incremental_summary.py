"""增量订单报表 - 汇总和格式模块

包含创建Excel汇总行和调整格式的逻辑。
"""

from openpyxl import Workbook
from openpyxl.styles import Font


def write_summary_row(
    ws_orders,
    row_idx: int,
    total_amount: float,
    total_interest: float,
    total_principal: float,
    order_count: int,
    styles: dict,
) -> None:
    """写入汇总行

    Args:
        ws_orders: Excel工作表对象
        row_idx: 行号
        total_amount: 总金额
        total_interest: 总利息
        total_principal: 总本金
        order_count: 订单数
        styles: 样式字典
    """
    if order_count > 0:
        summary_cell = ws_orders.cell(row=row_idx, column=1, value="汇总")
        summary_cell.font = Font(bold=True)
        summary_cell.border = styles["border"]
        ws_orders.cell(row=row_idx, column=2, value="-").border = styles["border"]
        ws_orders.cell(row=row_idx, column=3, value="-").border = styles["border"]
        total_amount_cell = ws_orders.cell(row=row_idx, column=4, value=total_amount)
        total_amount_cell.border = styles["border"]
        total_amount_cell.number_format = "#,##0.00"
        total_amount_cell.alignment = styles["right_align"]
        total_amount_cell.font = Font(bold=True)
        total_interest_cell = ws_orders.cell(
            row=row_idx, column=5, value=total_interest
        )
        total_interest_cell.border = styles["border"]
        total_interest_cell.number_format = "#,##0.00"
        total_interest_cell.alignment = styles["right_align"]
        total_interest_cell.font = Font(bold=True)
        total_principal_cell = ws_orders.cell(
            row=row_idx, column=6, value=total_principal
        )
        total_principal_cell.border = styles["border"]
        total_principal_cell.number_format = "#,##0.00"
        total_principal_cell.alignment = styles["right_align"]
        total_principal_cell.font = Font(bold=True)
        count_cell = ws_orders.cell(row=row_idx, column=7, value=f"{order_count}个订单")
        count_cell.font = Font(bold=True)
        count_cell.border = styles["border"]
        ws_orders.cell(row=row_idx, column=8, value="-").border = styles["border"]


def adjust_column_widths(ws_orders) -> None:
    """调整列宽

    Args:
        ws_orders: Excel工作表对象
    """
    ws_orders.column_dimensions["A"].width = 12
    ws_orders.column_dimensions["B"].width = 15
    ws_orders.column_dimensions["C"].width = 8
    ws_orders.column_dimensions["D"].width = 15
    ws_orders.column_dimensions["E"].width = 15
    ws_orders.column_dimensions["F"].width = 15
    ws_orders.column_dimensions["G"].width = 12
    ws_orders.column_dimensions["H"].width = 30
