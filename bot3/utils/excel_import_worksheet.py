"""Excel导入工作表处理模块

包含Excel工作表查找和表头解析的逻辑。
"""

import logging
from typing import Optional, Tuple

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


def find_worksheet(wb: Workbook) -> Tuple[Optional[Worksheet], bool]:
    """查找订单总表工作表

    Returns:
        tuple: (worksheet, found) - 工作表对象和是否找到
    """
    if "订单总表" in wb.sheetnames:
        return wb["订单总表"], True
    elif len(wb.sheetnames) > 0:
        return wb[wb.sheetnames[0]], True  # 使用第一个工作表
    else:
        logger.error("Excel文件中没有找到工作表")
        return None, False


def find_header_row(ws: Worksheet, max_rows: int = 10) -> Optional[int]:
    """查找表头行

    Args:
        ws: 工作表对象
        max_rows: 最大搜索行数

    Returns:
        int: 表头行索引（从1开始），如果未找到返回None
    """
    for row_idx in range(1, min(max_rows, ws.max_row + 1)):
        row_values = [str(cell.value or "").strip() for cell in ws[row_idx]]
        if "订单号" in row_values or "订单" in row_values:
            return row_idx
    return None
