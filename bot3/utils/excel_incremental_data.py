"""增量订单报表 - 数据行模块

包含创建Excel数据行的逻辑。
"""

from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font

from constants import ORDER_STATES


def _write_order_main_row(ws_orders, order: Dict, row_idx: int, styles: dict) -> None:
    """写入订单主行信息

    Args:
        ws_orders: Excel工作表对象
        order: 订单数据字典
        row_idx: 行号
        styles: 样式字典
    """
    date_str = order.get("date", "")[:10] if order.get("date") else "未知"
    order_id = order.get("order_id", "未知")
    customer = order.get("customer", "未知")
    amount = float(order.get("amount", 0) or 0)
    total_interest_amount = float(order.get("total_interest", 0) or 0)
    principal_reduction = float(order.get("principal_reduction", 0) or 0)
    state = ORDER_STATES.get(order.get("state", ""), order.get("state", "未知"))
    note = order.get("note", "")

    ws_orders.cell(row=row_idx, column=1, value=date_str).border = styles["border"]
    ws_orders.cell(row=row_idx, column=2, value=order_id).border = styles["border"]
    ws_orders.cell(row=row_idx, column=3, value=customer).border = styles["border"]
    amount_cell = ws_orders.cell(row=row_idx, column=4, value=amount)
    amount_cell.border = styles["border"]
    amount_cell.number_format = "#,##0.00"
    amount_cell.alignment = styles["right_align"]
    interest_cell = ws_orders.cell(row=row_idx, column=5, value=total_interest_amount)
    interest_cell.border = styles["border"]
    interest_cell.number_format = "#,##0.00"
    interest_cell.alignment = styles["right_align"]
    principal_cell = ws_orders.cell(row=row_idx, column=6, value=principal_reduction)
    principal_cell.border = styles["border"]
    principal_cell.number_format = "#,##0.00"
    principal_cell.alignment = styles["right_align"]
    ws_orders.cell(row=row_idx, column=7, value=state).border = styles["border"]
    ws_orders.cell(row=row_idx, column=8, value=note).border = styles["border"]


def _write_interest_details(
    ws_orders, order: Dict, row_idx: int, styles: dict, detail_font: Font
) -> int:
    """写入利息明细行并创建分组

    Args:
        ws_orders: Excel工作表对象
        order: 订单数据字典
        row_idx: 起始行号
        styles: 样式字典
        detail_font: 明细字体

    Returns:
        下一行号
    """
    interests = order.get("interests", [])
    if not interests:
        return row_idx

    detail_start_row = row_idx + 1
    for interest in interests:
        interest_date = (
            interest.get("date", "")[:10] if interest.get("date") else "未知"
        )
        interest_amount = float(interest.get("amount", 0) or 0)

        for col in range(1, 9):
            cell = ws_orders.cell(row=row_idx + 1, column=col, value="")
            cell.border = styles["border"]
            if col == 5:
                detail_text = f"  └─ {interest_date}: {interest_amount:,.2f}"
                cell.value = detail_text
                cell.font = detail_font
        row_idx += 1

    detail_end_row = row_idx
    for detail_row in range(detail_start_row, detail_end_row + 1):
        ws_orders.row_dimensions[detail_row].outline_level = 1
        ws_orders.row_dimensions[detail_row].hidden = True

    return row_idx


def write_order_data_rows(
    ws_orders, orders_data: List[Dict], styles: dict, start_row: int = 3
) -> tuple[int, float, float, float, int]:
    """写入订单数据行

    Args:
        ws_orders: Excel工作表对象
        orders_data: 订单数据列表
        styles: 样式字典
        start_row: 起始行号

    Returns:
        Tuple: (下一行号, 总金额, 总利息, 总本金, 订单数)
    """
    row_idx = start_row
    total_amount = 0.0
    total_interest = 0.0
    total_principal = 0.0
    order_count = 0
    detail_font = Font(size=10, color="666666")

    for order in orders_data:
        _write_order_main_row(ws_orders, order, row_idx, styles)
        row_idx = _write_interest_details(
            ws_orders, order, row_idx, styles, detail_font
        )
        row_idx += 1

        amount = float(order.get("amount", 0) or 0)
        total_interest_amount = float(order.get("total_interest", 0) or 0)
        principal_reduction = float(order.get("principal_reduction", 0) or 0)
        total_amount += amount
        total_interest += total_interest_amount
        total_principal += principal_reduction
        order_count += 1

    return row_idx, total_amount, total_interest, total_principal, order_count
