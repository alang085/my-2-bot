"""订单总表Excel - 汇总模块

包含创建订单总表汇总行和调整列宽的逻辑。
"""

from typing import List

from openpyxl.styles import Font


def create_orders_summary(
    ws_orders,
    orders: List,
    row_idx: int,
    total_amount: float,
    total_interest_all: float,
    styles: dict,
) -> None:
    """创建订单总表汇总行

    Args:
        ws_orders: 工作表对象
        orders: 订单列表
        row_idx: 汇总行号
        total_amount: 总金额
        total_interest_all: 总利息
        styles: 样式字典
    """
    # 汇总行
    if orders:
        ws_orders.merge_cells(f"A{row_idx}:E{row_idx}")
        summary_cell = ws_orders.cell(
            row=row_idx, column=1, value=f"总计: {len(orders)} 个订单"
        )
        summary_cell.font = Font(bold=True)
        summary_cell.alignment = styles["right_align"]
        summary_cell.border = styles["border"]
        total_amount_cell = ws_orders.cell(row=row_idx, column=6, value=total_amount)
        total_amount_cell.border = styles["border"]
        total_amount_cell.number_format = "#,##0.00"
        total_amount_cell.alignment = styles["right_align"]
        total_amount_cell.font = Font(bold=True)
        total_interest_cell = ws_orders.cell(
            row=row_idx, column=7, value=total_interest_all
        )
        total_interest_cell.border = styles["border"]
        total_interest_cell.number_format = "#,##0.00"
        total_interest_cell.alignment = styles["right_align"]
        total_interest_cell.font = Font(bold=True)
        ws_orders.cell(row=row_idx, column=8, value="").border = styles["border"]


def adjust_orders_column_width(ws_orders) -> None:
    """调整订单总表列宽

    Args:
        ws_orders: 工作表对象
    """
    ws_orders.column_dimensions["A"].width = 12
    ws_orders.column_dimensions["B"].width = 15
    ws_orders.column_dimensions["C"].width = 8
    ws_orders.column_dimensions["D"].width = 12
    ws_orders.column_dimensions["E"].width = 15
    ws_orders.column_dimensions["F"].width = 12
    ws_orders.column_dimensions["G"].width = 15
    ws_orders.column_dimensions["H"].width = 40
    # 隐藏利息历史列（H列，第8列）- 需要记录但可以隐藏
    ws_orders.column_dimensions["H"].hidden = True
