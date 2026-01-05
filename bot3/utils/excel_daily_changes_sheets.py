"""每日变化数据工作表创建模块

包含每日变化数据Excel中各个工作表的创建函数。
"""

# 标准库
from datetime import datetime, timedelta
from typing import Dict, List

# 第三方库
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

# 本地模块
import db_operations
from utils.excel_export import format_datetime_to_beijing


def _calculate_financial_totals(
    daily_summary: Dict, income_records: List[Dict]
) -> Dict[str, float]:
    """计算财务汇总数据

    Args:
        daily_summary: 日结数据
        income_records: 收入记录列表

    Returns:
        包含各种财务总计的字典
    """
    total_principal = sum(
        float(record.get("amount", 0) or 0)
        for record in income_records
        if record.get("type") == "principal_reduction"
    )
    total_interest = daily_summary.get("daily_interest", 0.0)
    # 违约完成金额按实际收入记录计算
    breach_end_income_amount = sum(
        float(record.get("amount", 0) or 0)
        for record in income_records
        if record.get("type") == "breach_end"
    )
    total_income = total_interest + total_principal + breach_end_income_amount
    total_expenses = daily_summary.get("company_expenses", 0.0) + daily_summary.get(
        "other_expenses", 0.0
    )
    net_income = total_income - total_expenses

    return {
        "total_principal": total_principal,
        "total_interest": total_interest,
        "total_income": total_income,
        "total_expenses": total_expenses,
        "net_income": net_income,
        "breach_end_income_amount": breach_end_income_amount,
    }


def _build_summary_data(
    daily_summary: Dict,
    financial_totals: Dict[str, float],
    new_orders: List[Dict] = None,
    income_records: List[Dict] = None,
) -> List[List]:
    """构建汇总数据列表

    Args:
        daily_summary: 日结数据
        financial_totals: 财务总计字典
        new_orders: 新订单列表（用于区分新老客户）
        income_records: 收入记录列表（用于计算本金归还订单数）

    Returns:
        汇总数据列表
    """
    completed_amount = daily_summary.get(
        "completed_orders_amount", daily_summary.get("completed_amount", 0.0)
    )
    # 违约完成金额现在从 financial_totals 中获取实际收入
    breach_end_amount = financial_totals.get("breach_end_income_amount", 0.0)

    # 区分新老订单
    new_client_orders = []
    old_client_orders = []
    if new_orders:
        new_client_orders = [
            order for order in new_orders if order.get("customer") == "A"
        ]
        old_client_orders = [
            order for order in new_orders if order.get("customer") == "B"
        ]

    new_client_count = len(new_client_orders)
    new_client_amount = sum(
        float(order.get("amount", 0) or 0) for order in new_client_orders
    )
    old_client_count = len(old_client_orders)
    old_client_amount = sum(
        float(order.get("amount", 0) or 0) for order in old_client_orders
    )

    # 本金归还订单数和金额
    principal_records = []
    if income_records:
        principal_records = [
            record
            for record in income_records
            if record.get("type") == "principal_reduction"
        ]
    principal_repayment_count = len(principal_records)
    principal_repayment_amount = financial_totals.get("total_principal", 0.0)

    return [
        ["新客户新增订单数", new_client_count],
        ["新客户新增订单金额", new_client_amount],
        ["老客户新增订单数", old_client_count],
        ["老客户新增订单金额", old_client_amount],
        ["完结订单数", daily_summary.get("completed_orders_count", 0)],
        ["完结订单金额", completed_amount],
        ["违约完成数", daily_summary.get("breach_end_orders_count", 0)],
        ["违约完成金额", breach_end_amount],
        ["当日利息", financial_totals.get("total_interest", 0.0)],
        ["本金归还订单数", principal_repayment_count],
        ["本金归还金额", principal_repayment_amount],
        # 移除"收入明细汇总"行
        ["公司开销", daily_summary.get("company_expenses", 0.0)],
        ["其他开销", daily_summary.get("other_expenses", 0.0)],
        ["总开销", financial_totals.get("total_expenses", 0.0)],
        ["净收入", financial_totals.get("net_income", 0.0)],
    ]


def _write_summary_data_to_sheet(
    ws_summary, summary_data: List[List], styles: Dict, start_row: int = 3
) -> None:
    """将汇总数据写入工作表

    Args:
        ws_summary: 工作表对象
        summary_data: 汇总数据列表
        styles: 样式字典
        start_row: 起始行号（默认为3）
    """
    row_idx = start_row
    for label, value in summary_data:
        label_cell = ws_summary.cell(row=row_idx, column=1, value=label)
        label_cell.font = Font(bold=True)
        label_cell.border = styles["border"]
        value_cell = ws_summary.cell(row=row_idx, column=2, value=value)
        if isinstance(value, float):
            value_cell.number_format = "#,##0.00"
            value_cell.alignment = styles["right_align"]
        else:
            value_cell.alignment = styles["center_align"]
        value_cell.border = styles["border"]
        row_idx += 1


async def _calculate_monthly_summary(
    date: str, new_orders: List[Dict], income_records: List[Dict]
) -> Dict:
    """计算月度汇总数据（通过累加每天的数据）

    Args:
        date: 当前日期（YYYY-MM-DD格式）
        new_orders: 当日新订单列表
        income_records: 当日收入记录列表

    Returns:
        月度汇总数据字典
    """
    try:
        # 计算月初日期（当月1日）
        date_obj = datetime.strptime(date, "%Y-%m-%d")
        month_start = date_obj.replace(day=1).strftime("%Y-%m-%d")

        # 获取当月所有日期的daily_summary并累加
        monthly_summary = {
            "new_orders_count": 0,
            "new_orders_amount": 0.0,
            "completed_orders_count": 0,
            "completed_orders_amount": 0.0,
            "breach_end_orders_count": 0,
            "breach_end_orders_amount": 0.0,
            "daily_interest": 0.0,
            "company_expenses": 0.0,
            "other_expenses": 0.0,
        }

        # 获取当月所有日期的daily_summary
        current_date_obj = datetime.strptime(month_start, "%Y-%m-%d")
        end_date_obj = datetime.strptime(date, "%Y-%m-%d")

        monthly_income_records = []
        monthly_new_orders = []

        while current_date_obj <= end_date_obj:
            current_date_str = current_date_obj.strftime("%Y-%m-%d")
            daily_data = await db_operations.get_daily_summary(current_date_str)
            if daily_data:
                monthly_summary["new_orders_count"] += daily_data.get(
                    "new_orders_count", 0
                )
                monthly_summary["new_orders_amount"] += daily_data.get(
                    "new_orders_amount", 0.0
                )
                monthly_summary["completed_orders_count"] += daily_data.get(
                    "completed_orders_count", 0
                )
                monthly_summary["completed_orders_amount"] += daily_data.get(
                    "completed_orders_amount", 0.0
                )
                monthly_summary["breach_end_orders_count"] += daily_data.get(
                    "breach_end_orders_count", 0
                )
                monthly_summary["breach_end_orders_amount"] += daily_data.get(
                    "breach_end_orders_amount", 0.0
                )
                monthly_summary["daily_interest"] += daily_data.get(
                    "daily_interest", 0.0
                )
                monthly_summary["company_expenses"] += daily_data.get(
                    "company_expenses", 0.0
                )
                monthly_summary["other_expenses"] += daily_data.get(
                    "other_expenses", 0.0
                )

            # 获取当日的收入记录和新订单
            day_income_records = await db_operations.get_income_records(
                current_date_str, current_date_str
            )
            monthly_income_records.extend(day_income_records)

            day_new_orders = await db_operations.get_new_orders_by_date(
                current_date_str
            )
            monthly_new_orders.extend(day_new_orders)

            current_date_obj += timedelta(days=1)

        # 计算月度财务总计
        monthly_financial_totals = _calculate_financial_totals(
            monthly_summary, monthly_income_records
        )

        # 构建月度汇总数据
        monthly_summary_data = _build_summary_data(
            monthly_summary,
            monthly_financial_totals,
            monthly_new_orders,
            monthly_income_records,
        )

        return {
            "monthly_summary_data": monthly_summary_data,
            "month": date_obj.strftime("%Y年%m月"),
        }
    except Exception as e:
        # 如果获取月度数据失败，返回空数据
        return {
            "monthly_summary_data": [],
            "month": "",
        }


def create_daily_summary_sheet_for_changes(
    wb: Workbook,
    date: str,
    daily_summary: Dict,
    income_records: List[Dict],
    styles: Dict,
    new_orders: List[Dict] = None,
    monthly_data: Dict = None,
) -> None:
    """创建每日变化数据汇总工作表

    Args:
        wb: Workbook对象
        date: 日期字符串
        daily_summary: 日结数据
        income_records: 收入记录列表
        styles: 样式字典
        new_orders: 新订单列表（可选）
        monthly_data: 月度汇总数据（可选）
    """
    ws_summary = wb.create_sheet("数据汇总", 0)
    ws_summary.merge_cells("A1:B1")
    ws_summary["A1"] = f"每日变化数据汇总 ({date})"
    ws_summary["A1"].font = styles["title_font"]
    ws_summary["A1"].alignment = styles["center_align"]

    financial_totals = _calculate_financial_totals(daily_summary, income_records)
    summary_data = _build_summary_data(
        daily_summary, financial_totals, new_orders, income_records
    )
    start_row = 3
    _write_summary_data_to_sheet(ws_summary, summary_data, styles, start_row)

    # 如果有月度数据，添加月度汇总
    if monthly_data and monthly_data.get("monthly_summary_data"):
        month = monthly_data.get("month", "")
        monthly_summary_data = monthly_data.get("monthly_summary_data", [])

        # 添加空行分隔
        current_row = start_row + len(summary_data) + 2

        # 添加月度汇总标题
        ws_summary.merge_cells(f"A{current_row}:B{current_row}")
        ws_summary.cell(
            row=current_row, column=1, value=f"本月数据汇总 ({month})"
        ).font = styles["title_font"]
        ws_summary.cell(row=current_row, column=1).alignment = styles["center_align"]

        # 写入月度汇总数据
        _write_summary_data_to_sheet(
            ws_summary, monthly_summary_data, styles, current_row + 1
        )

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 20


def create_new_orders_sheet_for_changes(
    wb: Workbook, date: str, new_orders: List[Dict], styles: Dict
) -> None:
    """创建新增订单工作表（每日变化）"""
    if not new_orders:
        return

    ws_new = wb.create_sheet("新增订单")
    ws_new.merge_cells("A1:D1")
    ws_new["A1"] = f"新增订单 ({date})"
    ws_new["A1"].font = styles["title_font"]
    ws_new["A1"].alignment = styles["center_align"]

    headers = ["时间", "订单号", "金额", "状态"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_new.cell(row=2, column=col_idx, value=header)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center_align"]
        cell.border = styles["border"]

    row_idx = 3
    total_amount = 0.0
    for order in new_orders:
        date_str = order.get("date", "")[:10] if order.get("date") else "未知"
        order_id = order.get("order_id", "未知")
        amount = float(order.get("amount", 0) or 0)
        total_amount += amount
        state = order.get("state", "未知")

        ws_new.cell(row=row_idx, column=1, value=date_str).border = styles["border"]
        ws_new.cell(row=row_idx, column=2, value=order_id).border = styles["border"]
        amount_cell = ws_new.cell(row=row_idx, column=3, value=amount)
        amount_cell.border = styles["border"]
        amount_cell.number_format = "#,##0.00"
        amount_cell.alignment = styles["right_align"]
        ws_new.cell(row=row_idx, column=4, value=state).border = styles["border"]
        row_idx += 1

    # 添加汇总行
    summary_row = row_idx + 1
    ws_new.merge_cells(f"A{summary_row}:B{summary_row}")
    ws_new.cell(
        row=summary_row, column=1, value=f"总计: {len(new_orders)} 个订单"
    ).font = Font(bold=True)
    ws_new.cell(row=summary_row, column=1).alignment = styles["right_align"]
    ws_new.cell(row=summary_row, column=1).border = styles["border"]
    total_amount_cell = ws_new.cell(row=summary_row, column=3, value=total_amount)
    total_amount_cell.font = Font(bold=True)
    total_amount_cell.number_format = "#,##0.00"
    total_amount_cell.alignment = styles["right_align"]
    total_amount_cell.border = styles["border"]
    ws_new.cell(row=summary_row, column=4, value="").border = styles["border"]

    ws_new.column_dimensions["A"].width = 12
    ws_new.column_dimensions["B"].width = 15
    ws_new.column_dimensions["C"].width = 15
    ws_new.column_dimensions["D"].width = 10


def create_completed_orders_sheet_for_changes(
    wb: Workbook, date: str, completed_orders: List[Dict], styles: Dict
) -> None:
    """创建完成订单工作表（每日变化）"""
    if not completed_orders:
        return

    ws_completed = wb.create_sheet("完成订单")
    ws_completed.merge_cells("A1:D1")
    ws_completed["A1"] = f"完成订单 ({date})"
    ws_completed["A1"].font = styles["title_font"]
    ws_completed["A1"].alignment = styles["center_align"]

    headers = ["时间", "订单号", "金额", "完成时间"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_completed.cell(row=2, column=col_idx, value=header)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center_align"]
        cell.border = styles["border"]

    row_idx = 3
    total_amount = 0.0
    for order in completed_orders:
        date_str = order.get("date", "")[:10] if order.get("date") else "未知"
        order_id = order.get("order_id", "未知")
        amount = float(order.get("amount", 0) or 0)
        total_amount += amount
        updated_at_raw = (
            order.get("updated_at", "") if order.get("updated_at") else "未知"
        )
        updated_at = (
            format_datetime_to_beijing(updated_at_raw)
            if updated_at_raw != "未知"
            else "未知"
        )

        ws_completed.cell(row=row_idx, column=1, value=date_str).border = styles[
            "border"
        ]
        ws_completed.cell(row=row_idx, column=2, value=order_id).border = styles[
            "border"
        ]
        amount_cell = ws_completed.cell(row=row_idx, column=3, value=amount)
        amount_cell.border = styles["border"]
        amount_cell.number_format = "#,##0.00"
        amount_cell.alignment = styles["right_align"]
        ws_completed.cell(row=row_idx, column=4, value=updated_at).border = styles[
            "border"
        ]
        row_idx += 1

    # 添加汇总行
    summary_row = row_idx + 1
    ws_completed.merge_cells(f"A{summary_row}:B{summary_row}")
    ws_completed.cell(
        row=summary_row, column=1, value=f"总计: {len(completed_orders)} 个订单"
    ).font = Font(bold=True)
    ws_completed.cell(row=summary_row, column=1).alignment = styles["right_align"]
    ws_completed.cell(row=summary_row, column=1).border = styles["border"]
    total_amount_cell = ws_completed.cell(row=summary_row, column=3, value=total_amount)
    total_amount_cell.font = Font(bold=True)
    total_amount_cell.number_format = "#,##0.00"
    total_amount_cell.alignment = styles["right_align"]
    total_amount_cell.border = styles["border"]
    ws_completed.cell(row=summary_row, column=4, value="").border = styles["border"]

    ws_completed.column_dimensions["A"].width = 12
    ws_completed.column_dimensions["B"].width = 15
    ws_completed.column_dimensions["C"].width = 15
    ws_completed.column_dimensions["D"].width = 20


def create_breach_end_orders_sheet_for_changes(
    wb: Workbook, date: str, breach_end_orders: List[Dict], styles: Dict
) -> None:
    """创建违约完成订单工作表（每日变化）"""
    if not breach_end_orders:
        return

    ws_breach = wb.create_sheet("违约完成订单")
    ws_breach.merge_cells("A1:D1")
    ws_breach["A1"] = f"违约完成订单 ({date})"
    ws_breach["A1"].font = styles["title_font"]
    ws_breach["A1"].alignment = styles["center_align"]

    headers = ["时间", "订单号", "金额", "完成时间"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_breach.cell(row=2, column=col_idx, value=header)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center_align"]
        cell.border = styles["border"]

    row_idx = 3
    total_amount = 0.0
    for order in breach_end_orders:
        date_str = order.get("date", "")[:10] if order.get("date") else "未知"
        order_id = order.get("order_id", "未知")
        amount = float(order.get("amount", 0) or 0)
        total_amount += amount
        updated_at_raw = (
            order.get("updated_at", "") if order.get("updated_at") else "未知"
        )
        updated_at = (
            format_datetime_to_beijing(updated_at_raw)
            if updated_at_raw != "未知"
            else "未知"
        )

        ws_breach.cell(row=row_idx, column=1, value=date_str).border = styles["border"]
        ws_breach.cell(row=row_idx, column=2, value=order_id).border = styles["border"]
        amount_cell = ws_breach.cell(row=row_idx, column=3, value=amount)
        amount_cell.border = styles["border"]
        amount_cell.number_format = "#,##0.00"
        amount_cell.alignment = styles["right_align"]
        ws_breach.cell(row=row_idx, column=4, value=updated_at).border = styles[
            "border"
        ]
        row_idx += 1

    # 添加汇总行
    summary_row = row_idx + 1
    ws_breach.merge_cells(f"A{summary_row}:B{summary_row}")
    ws_breach.cell(
        row=summary_row, column=1, value=f"总计: {len(breach_end_orders)} 个订单"
    ).font = Font(bold=True)
    ws_breach.cell(row=summary_row, column=1).alignment = styles["right_align"]
    ws_breach.cell(row=summary_row, column=1).border = styles["border"]
    total_amount_cell = ws_breach.cell(row=summary_row, column=3, value=total_amount)
    total_amount_cell.font = Font(bold=True)
    total_amount_cell.number_format = "#,##0.00"
    total_amount_cell.alignment = styles["right_align"]
    total_amount_cell.border = styles["border"]
    ws_breach.cell(row=summary_row, column=4, value="").border = styles["border"]

    ws_breach.column_dimensions["A"].width = 12
    ws_breach.column_dimensions["B"].width = 15
    ws_breach.column_dimensions["C"].width = 15
    ws_breach.column_dimensions["D"].width = 20


def _setup_income_sheet_header(ws_income: Worksheet, date: str, styles: Dict) -> None:
    """设置收入明细工作表标题和表头

    Args:
        ws_income: 工作表对象
        date: 日期
        styles: 样式字典
    """
    ws_income.merge_cells("A1:E1")
    ws_income["A1"] = f"收入明细 ({date})"
    ws_income["A1"].font = styles["title_font"]
    ws_income["A1"].alignment = styles["center_align"]

    headers = ["时间", "类型", "订单号", "金额", "备注"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_income.cell(row=2, column=col_idx, value=header)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center_align"]
        cell.border = styles["border"]


def _write_income_data_rows(
    ws_income: Worksheet, income_records: List[Dict], styles: Dict
) -> None:
    """写入收入数据行

    Args:
        ws_income: 工作表对象
        income_records: 收入记录列表
        styles: 样式字典
    """
    row_idx = 3
    type_map = {
        "completed": "订单完成",
        "breach_end": "违约完成",
        "interest": "利息收入",
        "principal_reduction": "本金减少",
    }

    for record in income_records:
        created_at_raw = (
            record.get("created_at", "") if record.get("created_at") else "未知"
        )
        created_at = (
            format_datetime_to_beijing(created_at_raw)
            if created_at_raw != "未知"
            else "未知"
        )
        income_type = record.get("type", "未知")
        type_name = type_map.get(income_type, income_type)
        order_id = record.get("order_id", "") or "全局"
        amount = record.get("amount", 0)
        note = record.get("note", "") or ""

        ws_income.cell(row=row_idx, column=1, value=created_at).border = styles[
            "border"
        ]
        ws_income.cell(row=row_idx, column=2, value=type_name).border = styles["border"]
        ws_income.cell(row=row_idx, column=3, value=order_id).border = styles["border"]
        amount_cell = ws_income.cell(
            row=row_idx, column=4, value=float(amount) if amount else 0
        )
        amount_cell.border = styles["border"]
        amount_cell.number_format = "#,##0.00"
        amount_cell.alignment = styles["right_align"]
        ws_income.cell(row=row_idx, column=5, value=note).border = styles["border"]
        row_idx += 1


def _set_income_sheet_column_widths(ws_income: Worksheet) -> None:
    """设置收入明细工作表列宽

    Args:
        ws_income: 工作表对象
    """
    ws_income.column_dimensions["A"].width = 20
    ws_income.column_dimensions["B"].width = 12
    ws_income.column_dimensions["C"].width = 15
    ws_income.column_dimensions["D"].width = 15
    ws_income.column_dimensions["E"].width = 30


def create_income_records_sheet(
    wb: Workbook, date: str, income_records: List[Dict], styles: Dict
) -> None:
    """创建收入明细工作表"""
    if not income_records:
        return

    ws_income = wb.create_sheet("收入明细")
    _setup_income_sheet_header(ws_income, date, styles)
    _write_income_data_rows(ws_income, income_records, styles)
    _set_income_sheet_column_widths(ws_income)


def create_expense_records_sheet(
    wb: Workbook, date: str, expense_records: List[Dict], styles: Dict
) -> None:
    """创建开销明细工作表"""
    if not expense_records:
        return

    ws_expense = wb.create_sheet("开销明细")
    ws_expense.merge_cells("A1:D1")
    ws_expense["A1"] = f"开销明细 ({date})"
    ws_expense["A1"].font = styles["title_font"]
    ws_expense["A1"].alignment = styles["center_align"]

    headers = ["时间", "类型", "金额", "备注"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_expense.cell(row=2, column=col_idx, value=header)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center_align"]
        cell.border = styles["border"]

    row_idx = 3
    type_map = {"company": "公司开销", "other": "其他开销"}
    for record in expense_records:
        date_str = record.get("date", "")[:10] if record.get("date") else "未知"
        expense_type = record.get("type", "未知")
        type_name = type_map.get(expense_type, expense_type)
        amount = record.get("amount", 0)
        note = record.get("note", "") or "无备注"

        ws_expense.cell(row=row_idx, column=1, value=date_str).border = styles["border"]
        ws_expense.cell(row=row_idx, column=2, value=type_name).border = styles[
            "border"
        ]
        amount_cell = ws_expense.cell(
            row=row_idx, column=3, value=float(amount) if amount else 0
        )
        amount_cell.border = styles["border"]
        amount_cell.number_format = "#,##0.00"
        amount_cell.alignment = styles["right_align"]
        ws_expense.cell(row=row_idx, column=4, value=note).border = styles["border"]
        row_idx += 1

    ws_expense.column_dimensions["A"].width = 12
    ws_expense.column_dimensions["B"].width = 12
    ws_expense.column_dimensions["C"].width = 15
    ws_expense.column_dimensions["D"].width = 40
