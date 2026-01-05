"""订单总表Excel - 数据行模块

包含写入订单总表数据行的逻辑。
"""

from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font

from utils.excel_orders_sheets import ORDER_STATES


def _calculate_order_interest(interests: List[Dict]) -> float:
    """计算订单总利息

    Args:
        interests: 利息记录列表

    Returns:
        总利息金额
    """
    return sum(float(interest.get("amount", 0) or 0) for interest in interests)


def _format_interest_details(interests: List[Dict]) -> str:
    """格式化利息明细

    Args:
        interests: 利息记录列表

    Returns:
        格式化后的利息明细文本
    """
    if not interests:
        return "无"

    interest_details = []
    for interest in interests:
        interest_date = (
            interest.get("date", "")[:10] if interest.get("date") else "未知"
        )
        interest_amount = float(interest.get("amount", 0) or 0)
        interest_details.append(f"{interest_date}: {interest_amount:,.2f}")
    return "\n".join(interest_details)


def _write_single_order_row(
    ws_orders, row_idx: int, order: Dict, styles: dict
) -> tuple[int, float, float]:
    """写入单个订单行

    Args:
        ws_orders: 工作表对象
        row_idx: 当前行号
        order: 订单字典
        styles: 样式字典

    Returns:
        (下一行号, 订单金额, 订单利息)
    """
    date_str = order.get("date", "")[:10] if order.get("date") else "未知"
    order_id = order.get("order_id", "未知")
    customer = order.get("customer", "未知")
    group_id = order.get("group_id", "未知")
    amount = order.get("amount", 0)
    state = ORDER_STATES.get(order.get("state", ""), order.get("state", "未知"))

    interests = order.get("interests", [])
    order_total_interest = _calculate_order_interest(interests)
    order_amount = float(amount) if amount else 0

    ws_orders.cell(row=row_idx, column=1, value=date_str).border = styles["border"]
    ws_orders.cell(row=row_idx, column=2, value=order_id).border = styles["border"]
    ws_orders.cell(row=row_idx, column=3, value=customer).border = styles["border"]
    ws_orders.cell(row=row_idx, column=4, value=group_id).border = styles["border"]
    amount_cell = ws_orders.cell(row=row_idx, column=5, value=order_amount)
    amount_cell.border = styles["border"]
    amount_cell.number_format = "#,##0.00"
    amount_cell.alignment = styles["right_align"]
    ws_orders.cell(row=row_idx, column=6, value=state).border = styles["border"]

    interest_cell = ws_orders.cell(row=row_idx, column=7, value=order_total_interest)
    interest_cell.border = styles["border"]
    interest_cell.number_format = "#,##0.00"
    interest_cell.alignment = styles["right_align"]

    interest_text = _format_interest_details(interests)
    ws_orders.cell(row=row_idx, column=8, value=interest_text).border = styles["border"]

    return row_idx + 1, order_amount, order_total_interest


def write_orders_data(
    ws_orders, orders: List[Dict], styles: dict, start_row: int = 3
) -> tuple[int, float, float]:
    """写入订单数据行

    Args:
        ws_orders: 工作表对象
        orders: 订单列表
        styles: 样式字典
        start_row: 起始行号

    Returns:
        Tuple: (下一行号, 总金额, 总利息)
    """
    row_idx = start_row
    total_amount = 0.0
    total_interest_all = 0.0

    for order in orders:
        row_idx, order_amount, order_interest = _write_single_order_row(
            ws_orders, row_idx, order, styles
        )
        total_amount += order_amount
        total_interest_all += order_interest

    return row_idx, total_amount, total_interest_all
