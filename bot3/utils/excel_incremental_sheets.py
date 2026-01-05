"""增量报表工作表创建模块

包含增量报表Excel中各个工作表的创建函数。
"""

# 标准库
from typing import Dict, List, Tuple

# 第三方库
from openpyxl import Workbook
from openpyxl.styles import Font

# 本地模块
from constants import ORDER_STATES


def create_incremental_orders_sheet(
    wb: Workbook,
    baseline_date: str,
    current_date: str,
    orders_data: List[Dict],
    styles: Dict,
) -> None:
    """创建增量订单报表工作表"""
    from utils.excel_incremental_data import write_order_data_rows
    from utils.excel_incremental_header import create_sheet_header
    from utils.excel_incremental_summary import (adjust_column_widths,
                                                 write_summary_row)

    ws_orders = wb.create_sheet("增量订单报表", 0)

    # 创建标题和表头
    create_sheet_header(ws_orders, baseline_date, current_date, styles)

    # 写入数据行
    row_idx, total_amount, total_interest, total_principal, order_count = (
        write_order_data_rows(ws_orders, orders_data, styles, start_row=3)
    )

    # 写入汇总行
    write_summary_row(
        ws_orders,
        row_idx,
        total_amount,
        total_interest,
        total_principal,
        order_count,
        styles,
    )

    # 调整列宽
    adjust_column_widths(ws_orders)


def _create_expense_sheet_header(ws_expense, baseline_date: str, styles: Dict) -> None:
    """创建开销明细表头

    Args:
        ws_expense: 工作表对象
        baseline_date: 基准日期
        styles: 样式字典
    """
    ws_expense.merge_cells("A1:D1")
    ws_expense["A1"] = f"开销明细 (基准日期: {baseline_date})"
    ws_expense["A1"].font = styles["title_font"]
    ws_expense["A1"].alignment = styles["center_align"]


def _write_expense_headers(ws_expense, styles: Dict) -> None:
    """写入开销表头行

    Args:
        ws_expense: 工作表对象
        styles: 样式字典
    """
    headers = ["日期", "类型", "金额", "备注"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_expense.cell(row=2, column=col_idx, value=header)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center_align"]
        cell.border = styles["border"]


def _write_expense_rows(
    ws_expense, expense_records: List[Dict], styles: Dict, start_row: int
) -> Tuple[int, float]:
    """写入开销数据行

    Args:
        ws_expense: 工作表对象
        expense_records: 开销记录列表
        styles: 样式字典
        start_row: 起始行号

    Returns:
        (下一行号, 总金额)
    """
    row_idx = start_row
    total_expense = 0.0
    type_map = {"company": "公司开销", "other": "其他开销"}

    for record in expense_records:
        date_str = record.get("date", "")[:10] if record.get("date") else "未知"
        expense_type = record.get("type", "未知")
        type_name = type_map.get(expense_type, expense_type)
        amount = float(record.get("amount", 0) or 0)
        note = record.get("note", "") or "无备注"

        ws_expense.cell(row=row_idx, column=1, value=date_str).border = styles["border"]
        ws_expense.cell(row=row_idx, column=2, value=type_name).border = styles[
            "border"
        ]
        amount_cell = ws_expense.cell(row=row_idx, column=3, value=amount)
        amount_cell.border = styles["border"]
        amount_cell.number_format = "#,##0.00"
        amount_cell.alignment = styles["right_align"]
        ws_expense.cell(row=row_idx, column=4, value=note).border = styles["border"]

        total_expense += amount
        row_idx += 1

    return row_idx, total_expense


def _write_expense_summary(
    ws_expense, total_expense: float, row_idx: int, styles: Dict
) -> None:
    """写入开销汇总行

    Args:
        ws_expense: 工作表对象
        total_expense: 总金额
        row_idx: 行号
        styles: 样式字典
    """
    summary_cell = ws_expense.cell(row=row_idx, column=1, value="汇总")
    summary_cell.font = Font(bold=True)
    summary_cell.border = styles["border"]
    ws_expense.cell(row=row_idx, column=2, value="-").border = styles["border"]
    total_cell = ws_expense.cell(row=row_idx, column=3, value=total_expense)
    total_cell.border = styles["border"]
    total_cell.number_format = "#,##0.00"
    total_cell.alignment = styles["right_align"]
    total_cell.font = Font(bold=True)
    ws_expense.cell(row=row_idx, column=4, value="-").border = styles["border"]


def _set_expense_column_widths(ws_expense) -> None:
    """设置开销列宽

    Args:
        ws_expense: 工作表对象
    """
    ws_expense.column_dimensions["A"].width = 12
    ws_expense.column_dimensions["B"].width = 12
    ws_expense.column_dimensions["C"].width = 15
    ws_expense.column_dimensions["D"].width = 40


def create_incremental_expense_sheet(
    wb: Workbook, baseline_date: str, expense_records: List[Dict], styles: Dict
) -> None:
    """创建增量开销明细工作表"""
    if not expense_records:
        return

    ws_expense = wb.create_sheet("开销明细")
    _create_expense_sheet_header(ws_expense, baseline_date, styles)
    _write_expense_headers(ws_expense, styles)

    row_idx, total_expense = _write_expense_rows(ws_expense, expense_records, styles, 3)

    if expense_records:
        _write_expense_summary(ws_expense, total_expense, row_idx, styles)

    _set_expense_column_widths(ws_expense)
