"""Excel导出工具"""

# 标准库
import logging
import os
from datetime import datetime
from typing import Dict, List

# 第三方库
from openpyxl import Workbook

# 本地模块
import db_operations
from utils.excel_export_helpers import (
    create_breach_end_orders_sheet, create_breach_end_orders_sheet_for_changes,
    create_breach_orders_sheet, create_chat_id_mapping_sheet,
    create_completed_orders_sheet, create_completed_orders_sheet_for_changes,
    create_daily_summary_sheet, create_daily_summary_sheet_for_changes,
    create_expense_records_sheet, create_income_records_sheet,
    create_incremental_expense_sheet, create_incremental_orders_sheet,
    create_new_orders_sheet_for_changes, create_orders_sheet, get_excel_styles)

logger = logging.getLogger(__name__)


def format_datetime_to_beijing(datetime_str: str) -> str:
    """将时间字符串转换为北京时间显示（使用统一的时区处理函数）"""
    if not datetime_str or datetime_str == "未知":
        return datetime_str

    # 如果是纯日期字符串（YYYY-MM-DD），直接返回
    if len(datetime_str) == 10 and datetime_str.count("-") == 2:
        return datetime_str

    # 使用统一的时区处理函数
    return datetime_str_to_beijing_str(datetime_str)


def create_excel_file(params: "ExcelFileParams") -> str:
    """创建Excel文件

    Args:
        params: Excel文件参数

    Returns:
        文件路径
    """
    from utils.excel_export_data import ExcelFileParams

    wb = Workbook()

    # 删除默认工作表
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # 获取样式配置
    styles = get_excel_styles()

    # 创建各个工作表
    create_orders_sheet(wb, params.orders, styles)
    create_completed_orders_sheet(wb, params.completed_orders or [], styles)
    create_breach_orders_sheet(wb, params.breach_orders or [], styles)
    create_breach_end_orders_sheet(wb, params.breach_end_orders or [], styles)
    create_daily_summary_sheet(wb, params.daily_summary, styles)
    create_chat_id_mapping_sheet(wb, params.all_orders_for_chat_id or [], styles)

    # 保存文件
    wb.save(params.file_path)
    return params.file_path


async def _fetch_orders_interests(order_ids: List[str]) -> dict:
    """批量获取订单利息记录

    Args:
        order_ids: 订单ID列表

    Returns:
        订单ID到利息记录的映射
    """
    interests_map = {}
    if not order_ids:
        return interests_map

    try:
        interests_map = await db_operations.get_interests_by_order_ids(order_ids)
        logger.debug(f"批量获取了 {len(order_ids)} 个订单的利息记录")
    except Exception as e:
        logger.error(f"批量获取利息记录失败: {e}", exc_info=True)
        for order_id in order_ids:
            try:
                interests = await db_operations.get_all_interest_by_order_id(order_id)
                interests_map[order_id] = interests
            except Exception as e2:
                logger.error(f"获取订单 {order_id} 的利息记录失败: {e2}")
                interests_map[order_id] = []

    return interests_map


def _attach_interests_to_orders(orders: List[Dict], interests_map: dict) -> List[Dict]:
    """为订单附加利息记录

    Args:
        orders: 订单列表
        interests_map: 利息记录映射

    Returns:
        附加了利息记录的订单列表
    """
    orders_with_interests = []
    for order in orders:
        order_id = order.get("order_id")
        order_copy = order.copy()
        order_copy["interests"] = interests_map.get(order_id, [])
        orders_with_interests.append(order_copy)
    return orders_with_interests


def _prepare_excel_file_path() -> str:
    """准备Excel文件路径

    Returns:
        文件路径
    """
    temp_dir = os.path.join(os.path.dirname(__file__), "..", "temp")
    os.makedirs(temp_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"订单报表_{timestamp}.xlsx"
    return os.path.join(temp_dir, file_name)


async def _fetch_all_orders_for_chat_id() -> List[Dict]:
    """获取所有订单用于chat_id对应表

    Returns:
        订单列表
    """
    try:
        all_orders = await db_operations.search_orders_advanced_all_states({})
        logger.debug(f"获取到 {len(all_orders)} 个订单用于chat_id对应表")
        return all_orders
    except Exception as e:
        logger.warning(f"获取所有订单失败: {e}")
        return []


async def export_orders_to_excel(
    orders: List[Dict],
    completed_orders: List[Dict] = None,
    breach_orders: List[Dict] = None,
    breach_end_orders: List[Dict] = None,
    daily_interest: float = 0,
    daily_summary: Dict = None,
) -> str:
    """导出订单到Excel文件（异步版本，优化批量查询）"""
    import asyncio

    order_ids = [order.get("order_id") for order in orders if order.get("order_id")]
    interests_map = await _fetch_orders_interests(order_ids)
    orders_with_interests = _attach_interests_to_orders(orders, interests_map)
    file_path = _prepare_excel_file_path()
    all_orders_for_chat_id = await _fetch_all_orders_for_chat_id()

    from utils.excel_export_data import ExcelFileParams

    excel_params = ExcelFileParams(
        file_path=file_path,
        orders=orders_with_interests,
        completed_orders=completed_orders,
        breach_orders=breach_orders,
        breach_end_orders=breach_end_orders,
        daily_interest=daily_interest,
        daily_summary=daily_summary,
        all_orders_for_chat_id=all_orders_for_chat_id,
    )

    loop = asyncio.get_event_loop()
    await loop.run_in_executor(None, create_excel_file, excel_params)

    return file_path


def create_daily_changes_excel_file(params: "DailyChangesExcelParams") -> str:
    """创建每日变化数据Excel文件

    Args:
        params: 每日变化Excel文件参数

    Returns:
        文件路径
    """
    from utils.excel_export_data import DailyChangesExcelParams

    wb = Workbook()

    # 删除默认工作表
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # 获取样式配置
    styles = get_excel_styles()

    # 创建各个工作表
    create_daily_summary_sheet_for_changes(
        wb,
        params.date,
        params.daily_summary,
        params.income_records,
        styles,
        params.new_orders,
        params.monthly_data,
    )
    create_new_orders_sheet_for_changes(wb, params.date, params.new_orders, styles)
    create_completed_orders_sheet_for_changes(
        wb, params.date, params.completed_orders, styles
    )
    create_breach_end_orders_sheet_for_changes(
        wb, params.date, params.breach_end_orders, styles
    )
    # create_income_records_sheet(wb, params.date, params.income_records, styles)  # 已删除收入明细工作表
    create_expense_records_sheet(wb, params.date, params.expense_records, styles)

    # 保存文件
    wb.save(params.file_path)
    return params.file_path


async def export_daily_changes_to_excel(date: str) -> str:
    """导出每日变化数据到Excel文件（异步版本）"""
    import asyncio

    # 获取所有数据
    new_orders = await db_operations.get_new_orders_by_date(date)
    completed_orders = await db_operations.get_completed_orders_by_date(date)
    breach_end_orders = await db_operations.get_breach_end_orders_by_date(date)
    income_records = await db_operations.get_income_records(date, date)
    expense_records = await db_operations.get_expense_records(date, date)

    # 计算汇总数据
    from utils.daily_report_generator import calculate_daily_summary

    daily_summary = await calculate_daily_summary(date)

    # 计算月度汇总数据
    from utils.excel_daily_changes_sheets import _calculate_monthly_summary

    monthly_data = await _calculate_monthly_summary(date, new_orders, income_records)

    # 创建临时文件
    temp_dir = os.path.join(os.path.dirname(__file__), "..", "temp")
    os.makedirs(temp_dir, exist_ok=True)

    file_name = f"每日变化数据_{date}.xlsx"
    file_path = os.path.join(temp_dir, file_name)

    from utils.excel_export_data import DailyChangesExcelParams

    changes_params = DailyChangesExcelParams(
        file_path=file_path,
        date=date,
        new_orders=new_orders,
        completed_orders=completed_orders,
        breach_end_orders=breach_end_orders,
        income_records=income_records,
        expense_records=expense_records,
        daily_summary=daily_summary,
        monthly_data=monthly_data,
    )

    # 在事件循环中运行同步函数
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(None, create_daily_changes_excel_file, changes_params)

    return file_path


def create_incremental_orders_report_file(
    file_path: str,
    baseline_date: str,
    current_date: str,
    orders_data: List[Dict],
    expense_records: List[Dict] = None,
) -> str:
    """创建增量订单报表Excel文件"""
    wb = Workbook()

    # 删除默认工作表
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # 获取样式配置
    styles = get_excel_styles()

    # 创建各个工作表
    create_incremental_orders_sheet(
        wb, baseline_date, current_date, orders_data, styles
    )
    create_incremental_expense_sheet(wb, baseline_date, expense_records or [], styles)

    # 保存文件
    wb.save(file_path)
    return file_path


async def export_incremental_orders_report_to_excel(
    baseline_date: str,
    current_date: str,
    orders_data: List[Dict],
    expense_records: List[Dict] = None,
) -> str:
    """导出增量订单报表到Excel文件（异步版本）"""
    import asyncio

    # 创建临时文件
    temp_dir = os.path.join(os.path.dirname(__file__), "..", "temp")
    os.makedirs(temp_dir, exist_ok=True)

    file_name = f"增量订单报表_{current_date}.xlsx"
    file_path = os.path.join(temp_dir, file_name)

    # 在事件循环中运行同步函数
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(
        None,
        create_incremental_orders_report_file,
        file_path,
        baseline_date,
        current_date,
        orders_data,
        expense_records,
    )

    return file_path
