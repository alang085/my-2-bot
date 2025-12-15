"""Excel导出工具"""
# 标准库
import logging
import os
from datetime import datetime
from typing import List, Dict

# 第三方库
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 本地模块
import db_operations
from constants import ORDER_STATES

logger = logging.getLogger(__name__)


def create_excel_file(file_path: str, orders: List[Dict], completed_orders: List[Dict] = None, 
                     breach_end_orders: List[Dict] = None, daily_interest: float = 0,
                     daily_summary: Dict = None) -> str:
    """创建Excel文件"""
    wb = Workbook()
    
    # 删除默认工作表
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 定义样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # 1. 订单总表工作表
    ws_orders = wb.create_sheet("订单总表", 0)
    
    # 标题
    ws_orders.merge_cells('A1:E1')
    ws_orders['A1'] = "订单总表（有效订单）"
    ws_orders['A1'].font = Font(bold=True, size=14)
    ws_orders['A1'].alignment = center_align
    
    # 表头
    headers = ['时间', '订单号', '金额', '状态', '利息记录']
    for col_idx, header in enumerate(headers, 1):
        cell = ws_orders.cell(row=2, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # 数据行
    row_idx = 3
    for order in orders:
        date_str = order.get('date', '')[:10] if order.get('date') else '未知'
        order_id = order.get('order_id', '未知')
        amount = order.get('amount', 0)
        state = ORDER_STATES.get(order.get('state', ''), order.get('state', '未知'))
        
        # 订单基本信息行
        ws_orders.cell(row=row_idx, column=1, value=date_str).border = border
        ws_orders.cell(row=row_idx, column=2, value=order_id).border = border
        ws_orders.cell(row=row_idx, column=3, value=float(amount) if amount else 0).border = border
        ws_orders.cell(row=row_idx, column=3).number_format = '#,##0.00'
        ws_orders.cell(row=row_idx, column=3).alignment = right_align
        ws_orders.cell(row=row_idx, column=4, value=state).border = border
        
        # 获取利息记录（从传入的订单数据中获取，如果订单有interests字段）
        interests = order.get('interests', [])
        
        if interests:
            interest_text = "\n".join([
                f"{interest.get('date', '')[:10] if interest.get('date') else '未知'}: "
                f"{float(interest.get('amount', 0)):,.2f}"
                for interest in interests
            ])
            ws_orders.cell(row=row_idx, column=5, value=interest_text).border = border
        else:
            ws_orders.cell(row=row_idx, column=5, value="无").border = border
        
        row_idx += 1
    
    # 汇总行
    if daily_interest > 0:
        ws_orders.merge_cells(f'A{row_idx}:D{row_idx}')
        ws_orders.cell(row=row_idx, column=1, value="当日利息汇总:").font = Font(bold=True)
        ws_orders.cell(row=row_idx, column=5, value=float(daily_interest)).number_format = '#,##0.00'
        ws_orders.cell(row=row_idx, column=5).font = Font(bold=True)
        ws_orders.cell(row=row_idx, column=5).alignment = right_align
    
    # 调整列宽
    ws_orders.column_dimensions['A'].width = 12
    ws_orders.column_dimensions['B'].width = 15
    ws_orders.column_dimensions['C'].width = 15
    ws_orders.column_dimensions['D'].width = 10
    ws_orders.column_dimensions['E'].width = 30
    
    # 2. 已完成订单工作表
    if completed_orders:
        ws_completed = wb.create_sheet("已完成订单")
        ws_completed.merge_cells('A1:D1')
        ws_completed['A1'] = "已完成订单（当日）"
        ws_completed['A1'].font = Font(bold=True, size=14)
        ws_completed['A1'].alignment = center_align
        
        headers = ['时间', '订单号', '金额', '完成时间']
        for col_idx, header in enumerate(headers, 1):
            cell = ws_completed.cell(row=2, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        row_idx = 3
        for order in completed_orders:
            date_str = order.get('date', '')[:10] if order.get('date') else '未知'
            order_id = order.get('order_id', '未知')
            amount = order.get('amount', 0)
            updated_at = order.get('updated_at', '')[:19] if order.get('updated_at') else '未知'
            
            ws_completed.cell(row=row_idx, column=1, value=date_str).border = border
            ws_completed.cell(row=row_idx, column=2, value=order_id).border = border
            ws_completed.cell(row=row_idx, column=3, value=float(amount) if amount else 0).border = border
            ws_completed.cell(row=row_idx, column=3).number_format = '#,##0.00'
            ws_completed.cell(row=row_idx, column=3).alignment = right_align
            ws_completed.cell(row=row_idx, column=4, value=updated_at).border = border
            
            row_idx += 1
        
        ws_completed.column_dimensions['A'].width = 12
        ws_completed.column_dimensions['B'].width = 15
        ws_completed.column_dimensions['C'].width = 15
        ws_completed.column_dimensions['D'].width = 20
    
    # 3. 违约完成订单工作表
    if breach_end_orders:
        ws_breach = wb.create_sheet("违约完成订单")
        ws_breach.merge_cells('A1:D1')
        ws_breach['A1'] = "违约完成订单（当日有变动）"
        ws_breach['A1'].font = Font(bold=True, size=14)
        ws_breach['A1'].alignment = center_align
        
        headers = ['时间', '订单号', '金额', '完成时间']
        for col_idx, header in enumerate(headers, 1):
            cell = ws_breach.cell(row=2, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        row_idx = 3
        for order in breach_end_orders:
            date_str = order.get('date', '')[:10] if order.get('date') else '未知'
            order_id = order.get('order_id', '未知')
            amount = order.get('amount', 0)
            updated_at = order.get('updated_at', '')[:19] if order.get('updated_at') else '未知'
            
            ws_breach.cell(row=row_idx, column=1, value=date_str).border = border
            ws_breach.cell(row=row_idx, column=2, value=order_id).border = border
            ws_breach.cell(row=row_idx, column=3, value=float(amount) if amount else 0).border = border
            ws_breach.cell(row=row_idx, column=3).number_format = '#,##0.00'
            ws_breach.cell(row=row_idx, column=3).alignment = right_align
            ws_breach.cell(row=row_idx, column=4, value=updated_at).border = border
            
            row_idx += 1
        
        ws_breach.column_dimensions['A'].width = 12
        ws_breach.column_dimensions['B'].width = 15
        ws_breach.column_dimensions['C'].width = 15
        ws_breach.column_dimensions['D'].width = 20
    
    # 4. 日切数据汇总工作表
    if daily_summary:
        ws_summary = wb.create_sheet("日切数据汇总")
        ws_summary.merge_cells('A1:B1')
        ws_summary['A1'] = "日切数据汇总"
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary['A1'].alignment = center_align
        
        summary_data = [
            ['新增订单数', daily_summary.get('new_orders_count', 0)],
            ['新增订单金额', daily_summary.get('new_orders_amount', 0.0)],
            ['完结订单数', daily_summary.get('completed_orders_count', 0)],
            ['完结订单金额', daily_summary.get('completed_orders_amount', 0.0)],
            ['违约完成数', daily_summary.get('breach_end_orders_count', 0)],
            ['违约完成金额', daily_summary.get('breach_end_orders_amount', 0.0)],
            ['当日利息', daily_summary.get('daily_interest', 0.0)],
            ['公司开销', daily_summary.get('company_expenses', 0.0)],
            ['其他开销', daily_summary.get('other_expenses', 0.0)],
            ['总开销', daily_summary.get('company_expenses', 0.0) + daily_summary.get('other_expenses', 0.0)],
        ]
        
        row_idx = 3
        for label, value in summary_data:
            ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws_summary.cell(row=row_idx, column=1).border = border
            if isinstance(value, float):
                ws_summary.cell(row=row_idx, column=2, value=value).number_format = '#,##0.00'
                ws_summary.cell(row=row_idx, column=2).alignment = right_align
            else:
                ws_summary.cell(row=row_idx, column=2, value=value).alignment = center_align
            ws_summary.cell(row=row_idx, column=2).border = border
            row_idx += 1
        
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 20
    
    # 保存文件
    wb.save(file_path)
    return file_path


async def export_orders_to_excel(orders: List[Dict], completed_orders: List[Dict] = None,
                                breach_end_orders: List[Dict] = None, daily_interest: float = 0,
                                daily_summary: Dict = None) -> str:
    """导出订单到Excel文件（异步版本）"""
    import asyncio
    import tempfile
    
    # 为每个订单获取利息记录
    orders_with_interests = []
    for order in orders:
        order_id = order.get('order_id')
        if order_id:
            try:
                interests = await db_operations.get_all_interest_by_order_id(order_id)
                order_copy = order.copy()
                order_copy['interests'] = interests
                orders_with_interests.append(order_copy)
            except Exception as e:
                logger.error(f"获取订单 {order_id} 的利息记录失败: {e}")
                order_copy = order.copy()
                order_copy['interests'] = []
                orders_with_interests.append(order_copy)
        else:
            orders_with_interests.append(order)
    
    # 创建临时文件
    temp_dir = os.path.join(os.path.dirname(__file__), '..', 'temp')
    os.makedirs(temp_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"订单报表_{timestamp}.xlsx"
    file_path = os.path.join(temp_dir, file_name)
    
    # 在事件循环中运行同步函数
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(
        None,
        create_excel_file,
        file_path, orders_with_interests, completed_orders, breach_end_orders, daily_interest, daily_summary
    )
    
    return file_path


def create_daily_changes_excel_file(file_path: str, date: str, new_orders: List[Dict], 
                                   completed_orders: List[Dict], breach_end_orders: List[Dict],
                                   income_records: List[Dict], expense_records: List[Dict],
                                   daily_summary: Dict) -> str:
    """创建每日变化数据Excel文件"""
    wb = Workbook()
    
    # 删除默认工作表
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 定义样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    title_font = Font(bold=True, size=14)
    
    # 1. 数据汇总工作表
    ws_summary = wb.create_sheet("数据汇总", 0)
    ws_summary.merge_cells('A1:B1')
    ws_summary['A1'] = f"每日变化数据汇总 ({date})"
    ws_summary['A1'].font = title_font
    ws_summary['A1'].alignment = center_align
    
    summary_data = [
        ['新增订单数', daily_summary.get('new_orders_count', 0)],
        ['新增订单金额', daily_summary.get('new_orders_amount', 0.0)],
        ['完结订单数', daily_summary.get('completed_orders_count', 0)],
        ['完结订单金额', daily_summary.get('completed_amount', 0.0)],
        ['违约完成数', daily_summary.get('breach_end_orders_count', 0)],
        ['违约完成金额', daily_summary.get('breach_end_amount', 0.0)],
        ['当日利息', daily_summary.get('daily_interest', 0.0)],
        ['公司开销', daily_summary.get('company_expenses', 0.0)],
        ['其他开销', daily_summary.get('other_expenses', 0.0)],
        ['总开销', daily_summary.get('company_expenses', 0.0) + daily_summary.get('other_expenses', 0.0)],
    ]
    
    row_idx = 3
    for label, value in summary_data:
        ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=row_idx, column=1).border = border
        if isinstance(value, float):
            ws_summary.cell(row=row_idx, column=2, value=value).number_format = '#,##0.00'
            ws_summary.cell(row=row_idx, column=2).alignment = right_align
        else:
            ws_summary.cell(row=row_idx, column=2, value=value).alignment = center_align
        ws_summary.cell(row=row_idx, column=2).border = border
        row_idx += 1
    
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 20
    
    # 2. 新增订单工作表
    if new_orders:
        ws_new = wb.create_sheet("新增订单")
        ws_new.merge_cells('A1:D1')
        ws_new['A1'] = f"新增订单 ({date})"
        ws_new['A1'].font = title_font
        ws_new['A1'].alignment = center_align
        
        headers = ['时间', '订单号', '金额', '状态']
        for col_idx, header in enumerate(headers, 1):
            cell = ws_new.cell(row=2, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        row_idx = 3
        for order in new_orders:
            date_str = order.get('date', '')[:10] if order.get('date') else '未知'
            order_id = order.get('order_id', '未知')
            amount = order.get('amount', 0)
            state = order.get('state', '未知')
            
            ws_new.cell(row=row_idx, column=1, value=date_str).border = border
            ws_new.cell(row=row_idx, column=2, value=order_id).border = border
            ws_new.cell(row=row_idx, column=3, value=float(amount) if amount else 0).border = border
            ws_new.cell(row=row_idx, column=3).number_format = '#,##0.00'
            ws_new.cell(row=row_idx, column=3).alignment = right_align
            ws_new.cell(row=row_idx, column=4, value=state).border = border
            row_idx += 1
        
        ws_new.column_dimensions['A'].width = 12
        ws_new.column_dimensions['B'].width = 15
        ws_new.column_dimensions['C'].width = 15
        ws_new.column_dimensions['D'].width = 10
    
    # 3. 完成订单工作表
    if completed_orders:
        ws_completed = wb.create_sheet("完成订单")
        ws_completed.merge_cells('A1:D1')
        ws_completed['A1'] = f"完成订单 ({date})"
        ws_completed['A1'].font = title_font
        ws_completed['A1'].alignment = center_align
        
        headers = ['时间', '订单号', '金额', '完成时间']
        for col_idx, header in enumerate(headers, 1):
            cell = ws_completed.cell(row=2, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        row_idx = 3
        for order in completed_orders:
            date_str = order.get('date', '')[:10] if order.get('date') else '未知'
            order_id = order.get('order_id', '未知')
            amount = order.get('amount', 0)
            updated_at = order.get('updated_at', '')[:19] if order.get('updated_at') else '未知'
            
            ws_completed.cell(row=row_idx, column=1, value=date_str).border = border
            ws_completed.cell(row=row_idx, column=2, value=order_id).border = border
            ws_completed.cell(row=row_idx, column=3, value=float(amount) if amount else 0).border = border
            ws_completed.cell(row=row_idx, column=3).number_format = '#,##0.00'
            ws_completed.cell(row=row_idx, column=3).alignment = right_align
            ws_completed.cell(row=row_idx, column=4, value=updated_at).border = border
            row_idx += 1
        
        ws_completed.column_dimensions['A'].width = 12
        ws_completed.column_dimensions['B'].width = 15
        ws_completed.column_dimensions['C'].width = 15
        ws_completed.column_dimensions['D'].width = 20
    
    # 4. 违约完成订单工作表
    if breach_end_orders:
        ws_breach = wb.create_sheet("违约完成订单")
        ws_breach.merge_cells('A1:D1')
        ws_breach['A1'] = f"违约完成订单 ({date})"
        ws_breach['A1'].font = title_font
        ws_breach['A1'].alignment = center_align
        
        headers = ['时间', '订单号', '金额', '完成时间']
        for col_idx, header in enumerate(headers, 1):
            cell = ws_breach.cell(row=2, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        row_idx = 3
        for order in breach_end_orders:
            date_str = order.get('date', '')[:10] if order.get('date') else '未知'
            order_id = order.get('order_id', '未知')
            amount = order.get('amount', 0)
            updated_at = order.get('updated_at', '')[:19] if order.get('updated_at') else '未知'
            
            ws_breach.cell(row=row_idx, column=1, value=date_str).border = border
            ws_breach.cell(row=row_idx, column=2, value=order_id).border = border
            ws_breach.cell(row=row_idx, column=3, value=float(amount) if amount else 0).border = border
            ws_breach.cell(row=row_idx, column=3).number_format = '#,##0.00'
            ws_breach.cell(row=row_idx, column=3).alignment = right_align
            ws_breach.cell(row=row_idx, column=4, value=updated_at).border = border
            row_idx += 1
        
        ws_breach.column_dimensions['A'].width = 12
        ws_breach.column_dimensions['B'].width = 15
        ws_breach.column_dimensions['C'].width = 15
        ws_breach.column_dimensions['D'].width = 20
    
    # 5. 收入明细工作表（利息等）
    if income_records:
        ws_income = wb.create_sheet("收入明细")
        ws_income.merge_cells('A1:E1')
        ws_income['A1'] = f"收入明细 ({date})"
        ws_income['A1'].font = title_font
        ws_income['A1'].alignment = center_align
        
        headers = ['时间', '类型', '订单号', '金额', '备注']
        for col_idx, header in enumerate(headers, 1):
            cell = ws_income.cell(row=2, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        row_idx = 3
        type_map = {
            'completed': '订单完成',
            'breach_end': '违约完成',
            'interest': '利息收入',
            'principal_reduction': '本金减少'
        }
        for record in income_records:
            created_at = record.get('created_at', '')[:19] if record.get('created_at') else '未知'
            income_type = record.get('type', '未知')
            type_name = type_map.get(income_type, income_type)
            order_id = record.get('order_id', '') or '全局'
            amount = record.get('amount', 0)
            note = record.get('note', '') or ''
            
            ws_income.cell(row=row_idx, column=1, value=created_at).border = border
            ws_income.cell(row=row_idx, column=2, value=type_name).border = border
            ws_income.cell(row=row_idx, column=3, value=order_id).border = border
            ws_income.cell(row=row_idx, column=4, value=float(amount) if amount else 0).border = border
            ws_income.cell(row=row_idx, column=4).number_format = '#,##0.00'
            ws_income.cell(row=row_idx, column=4).alignment = right_align
            ws_income.cell(row=row_idx, column=5, value=note).border = border
            row_idx += 1
        
        ws_income.column_dimensions['A'].width = 20
        ws_income.column_dimensions['B'].width = 12
        ws_income.column_dimensions['C'].width = 15
        ws_income.column_dimensions['D'].width = 15
        ws_income.column_dimensions['E'].width = 30
    
    # 6. 开销明细工作表
    if expense_records:
        ws_expense = wb.create_sheet("开销明细")
        ws_expense.merge_cells('A1:D1')
        ws_expense['A1'] = f"开销明细 ({date})"
        ws_expense['A1'].font = title_font
        ws_expense['A1'].alignment = center_align
        
        headers = ['时间', '类型', '金额', '备注']
        for col_idx, header in enumerate(headers, 1):
            cell = ws_expense.cell(row=2, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        row_idx = 3
        type_map = {
            'company': '公司开销',
            'other': '其他开销'
        }
        for record in expense_records:
            date_str = record.get('date', '')[:10] if record.get('date') else '未知'
            expense_type = record.get('type', '未知')
            type_name = type_map.get(expense_type, expense_type)
            amount = record.get('amount', 0)
            note = record.get('note', '') or '无备注'
            
            ws_expense.cell(row=row_idx, column=1, value=date_str).border = border
            ws_expense.cell(row=row_idx, column=2, value=type_name).border = border
            ws_expense.cell(row=row_idx, column=3, value=float(amount) if amount else 0).border = border
            ws_expense.cell(row=row_idx, column=3).number_format = '#,##0.00'
            ws_expense.cell(row=row_idx, column=3).alignment = right_align
            ws_expense.cell(row=row_idx, column=4, value=note).border = border
            row_idx += 1
        
        ws_expense.column_dimensions['A'].width = 12
        ws_expense.column_dimensions['B'].width = 12
        ws_expense.column_dimensions['C'].width = 15
        ws_expense.column_dimensions['D'].width = 40
    
    # 保存文件
    wb.save(file_path)
    return file_path


async def export_daily_changes_to_excel(date: str) -> str:
    """导出每日变化数据到Excel文件（异步版本）"""
    import asyncio
    
    # 获取所有数据
    new_orders = await db_operations.get_new_orders_by_date(date)
    completed_orders = await db_operations.get_completed_orders_by_date(date)
    breach_end_orders = await db_operations.get_breach_end_orders_by_date(date)
    income_records = await db_operations.get_income_records(date, date)
    expense_records = await db_operations.get_expense_records(date, date)
    
    # 计算汇总数据
    from utils.daily_report_generator import calculate_daily_summary
    daily_summary = await calculate_daily_summary(date)
    
    # 创建临时文件
    temp_dir = os.path.join(os.path.dirname(__file__), '..', 'temp')
    os.makedirs(temp_dir, exist_ok=True)
    
    file_name = f"每日变化数据_{date}.xlsx"
    file_path = os.path.join(temp_dir, file_name)
    
    # 在事件循环中运行同步函数
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(
        None,
        create_daily_changes_excel_file,
        file_path, date, new_orders, completed_orders, breach_end_orders,
        income_records, expense_records, daily_summary
    )
    
    return file_path


def create_incremental_orders_report_file(file_path: str, baseline_date: str, current_date: str,
                                         orders_data: List[Dict], expense_records: List[Dict] = None) -> str:
    """创建增量订单报表Excel文件"""
    wb = Workbook()
    
    # 删除默认工作表
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 定义样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    title_font = Font(bold=True, size=14)
    detail_font = Font(size=10, color="666666")
    
    # 1. 增量订单报表工作表
    ws_orders = wb.create_sheet("增量订单报表", 0)
    
    # 标题
    ws_orders.merge_cells('A1:H1')
    ws_orders['A1'] = f"增量订单报表 (基准日期: {baseline_date}, 当前日期: {current_date})"
    ws_orders['A1'].font = title_font
    ws_orders['A1'].alignment = center_align
    
    # 表头
    headers = ['日期', '订单号', '会员', '订单金额', '利息总数', '归还本金', '订单状态', '备注']
    for col_idx, header in enumerate(headers, 1):
        cell = ws_orders.cell(row=2, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # 数据行
    row_idx = 3
    total_amount = 0.0
    total_interest = 0.0
    total_principal = 0.0
    order_count = 0
    
    for order in orders_data:
        date_str = order.get('date', '')[:10] if order.get('date') else '未知'
        order_id = order.get('order_id', '未知')
        customer = order.get('customer', '未知')
        amount = float(order.get('amount', 0) or 0)
        total_interest_amount = float(order.get('total_interest', 0) or 0)
        principal_reduction = float(order.get('principal_reduction', 0) or 0)
        state = ORDER_STATES.get(order.get('state', ''), order.get('state', '未知'))
        note = order.get('note', '')
        
        # 主行：订单信息
        ws_orders.cell(row=row_idx, column=1, value=date_str).border = border
        ws_orders.cell(row=row_idx, column=2, value=order_id).border = border
        ws_orders.cell(row=row_idx, column=3, value=customer).border = border
        ws_orders.cell(row=row_idx, column=4, value=amount).border = border
        ws_orders.cell(row=row_idx, column=4).number_format = '#,##0.00'
        ws_orders.cell(row=row_idx, column=4).alignment = right_align
        ws_orders.cell(row=row_idx, column=5, value=total_interest_amount).border = border
        ws_orders.cell(row=row_idx, column=5).number_format = '#,##0.00'
        ws_orders.cell(row=row_idx, column=5).alignment = right_align
        ws_orders.cell(row=row_idx, column=6, value=principal_reduction).border = border
        ws_orders.cell(row=row_idx, column=6).number_format = '#,##0.00'
        ws_orders.cell(row=row_idx, column=6).alignment = right_align
        ws_orders.cell(row=row_idx, column=7, value=state).border = border
        ws_orders.cell(row=row_idx, column=8, value=note).border = border
        
        main_row_idx = row_idx
        
        # 添加利息明细行（使用分组功能，默认隐藏）
        interests = order.get('interests', [])
        if interests:
            detail_start_row = row_idx + 1
            for interest in interests:
                interest_date = interest.get('date', '')[:10] if interest.get('date') else '未知'
                interest_amount = float(interest.get('amount', 0) or 0)
                
                # 明细行（缩进显示）
                ws_orders.cell(row=row_idx + 1, column=1, value='').border = border
                ws_orders.cell(row=row_idx + 1, column=2, value='').border = border
                ws_orders.cell(row=row_idx + 1, column=3, value='').border = border
                ws_orders.cell(row=row_idx + 1, column=4, value='').border = border
                detail_cell = ws_orders.cell(row=row_idx + 1, column=5, value=f"  └─ {interest_date}: {interest_amount:,.2f}")
                detail_cell.font = detail_font
                detail_cell.border = border
                ws_orders.cell(row=row_idx + 1, column=6, value='').border = border
                ws_orders.cell(row=row_idx + 1, column=7, value='').border = border
                ws_orders.cell(row=row_idx + 1, column=8, value='').border = border
                row_idx += 1
            
            detail_end_row = row_idx
            
            # 创建分组（可折叠，默认隐藏）
            # 使用outline_level属性创建分组
            for detail_row in range(detail_start_row, detail_end_row + 1):
                ws_orders.row_dimensions[detail_row].outline_level = 1
                ws_orders.row_dimensions[detail_row].hidden = True
        
        row_idx += 1
        
        # 累计汇总
        total_amount += amount
        total_interest += total_interest_amount
        total_principal += principal_reduction
        order_count += 1
    
    # 汇总行
    if order_count > 0:
        ws_orders.cell(row=row_idx, column=1, value="汇总").font = Font(bold=True)
        ws_orders.cell(row=row_idx, column=1).border = border
        ws_orders.cell(row=row_idx, column=2, value="-").border = border
        ws_orders.cell(row=row_idx, column=3, value="-").border = border
        ws_orders.cell(row=row_idx, column=4, value=total_amount).border = border
        ws_orders.cell(row=row_idx, column=4).number_format = '#,##0.00'
        ws_orders.cell(row=row_idx, column=4).alignment = right_align
        ws_orders.cell(row=row_idx, column=4).font = Font(bold=True)
        ws_orders.cell(row=row_idx, column=5, value=total_interest).border = border
        ws_orders.cell(row=row_idx, column=5).number_format = '#,##0.00'
        ws_orders.cell(row=row_idx, column=5).alignment = right_align
        ws_orders.cell(row=row_idx, column=5).font = Font(bold=True)
        ws_orders.cell(row=row_idx, column=6, value=total_principal).border = border
        ws_orders.cell(row=row_idx, column=6).number_format = '#,##0.00'
        ws_orders.cell(row=row_idx, column=6).alignment = right_align
        ws_orders.cell(row=row_idx, column=6).font = Font(bold=True)
        ws_orders.cell(row=row_idx, column=7, value=f"{order_count}个订单").border = border
        ws_orders.cell(row=row_idx, column=7).font = Font(bold=True)
        ws_orders.cell(row=row_idx, column=8, value="-").border = border
    
    # 调整列宽
    ws_orders.column_dimensions['A'].width = 12
    ws_orders.column_dimensions['B'].width = 15
    ws_orders.column_dimensions['C'].width = 8
    ws_orders.column_dimensions['D'].width = 15
    ws_orders.column_dimensions['E'].width = 20
    ws_orders.column_dimensions['F'].width = 15
    ws_orders.column_dimensions['G'].width = 12
    ws_orders.column_dimensions['H'].width = 30
    
    # 2. 开销明细工作表（如果有开销）
    if expense_records:
        ws_expense = wb.create_sheet("开销明细")
        ws_expense.merge_cells('A1:D1')
        ws_expense['A1'] = f"开销明细 (基准日期: {baseline_date})"
        ws_expense['A1'].font = title_font
        ws_expense['A1'].alignment = center_align
        
        headers = ['日期', '类型', '金额', '备注']
        for col_idx, header in enumerate(headers, 1):
            cell = ws_expense.cell(row=2, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        row_idx = 3
        total_expense = 0.0
        type_map = {
            'company': '公司开销',
            'other': '其他开销'
        }
        
        for record in expense_records:
            date_str = record.get('date', '')[:10] if record.get('date') else '未知'
            expense_type = record.get('type', '未知')
            type_name = type_map.get(expense_type, expense_type)
            amount = float(record.get('amount', 0) or 0)
            note = record.get('note', '') or '无备注'
            
            ws_expense.cell(row=row_idx, column=1, value=date_str).border = border
            ws_expense.cell(row=row_idx, column=2, value=type_name).border = border
            ws_expense.cell(row=row_idx, column=3, value=amount).border = border
            ws_expense.cell(row=row_idx, column=3).number_format = '#,##0.00'
            ws_expense.cell(row=row_idx, column=3).alignment = right_align
            ws_expense.cell(row=row_idx, column=4, value=note).border = border
            
            total_expense += amount
            row_idx += 1
        
        # 汇总行
        if expense_records:
            ws_expense.cell(row=row_idx, column=1, value="汇总").font = Font(bold=True)
            ws_expense.cell(row=row_idx, column=1).border = border
            ws_expense.cell(row=row_idx, column=2, value="-").border = border
            ws_expense.cell(row=row_idx, column=3, value=total_expense).border = border
            ws_expense.cell(row=row_idx, column=3).number_format = '#,##0.00'
            ws_expense.cell(row=row_idx, column=3).alignment = right_align
            ws_expense.cell(row=row_idx, column=3).font = Font(bold=True)
            ws_expense.cell(row=row_idx, column=4, value="-").border = border
        
        ws_expense.column_dimensions['A'].width = 12
        ws_expense.column_dimensions['B'].width = 12
        ws_expense.column_dimensions['C'].width = 15
        ws_expense.column_dimensions['D'].width = 40
    
    # 保存文件
    wb.save(file_path)
    return file_path


async def export_incremental_orders_report_to_excel(baseline_date: str, current_date: str,
                                                    orders_data: List[Dict],
                                                    expense_records: List[Dict] = None) -> str:
    """导出增量订单报表到Excel文件（异步版本）"""
    import asyncio
    
    # 创建临时文件
    temp_dir = os.path.join(os.path.dirname(__file__), '..', 'temp')
    os.makedirs(temp_dir, exist_ok=True)
    
    file_name = f"增量订单报表_{current_date}.xlsx"
    file_path = os.path.join(temp_dir, file_name)
    
    # 在事件循环中运行同步函数
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(
        None,
        create_incremental_orders_report_file,
        file_path, baseline_date, current_date, orders_data, expense_records
    )
    
    return file_path