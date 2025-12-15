"""完整测试所有Excel报表功能"""
import asyncio
import sys
import os
from pathlib import Path
from datetime import datetime, timedelta
import pytz

# 添加项目根目录到路径
project_root = Path(__file__).parent.absolute()
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

# 导入模块
import init_db
import db_operations
from utils.incremental_report_generator import get_or_create_baseline_date, prepare_incremental_data
from utils.excel_export import (
    export_incremental_orders_report_to_excel,
    export_daily_changes_to_excel,
    export_orders_to_excel
)
from utils.daily_report_generator import calculate_daily_summary
from utils.date_helpers import get_daily_period_date

BEIJING_TZ = pytz.timezone('Asia/Shanghai')


async def test_incremental_report_excel():
    """测试增量订单报表Excel"""
    print("=" * 60)
    print("测试1: 增量订单报表Excel")
    print("=" * 60)
    
    try:
        baseline_date = await get_or_create_baseline_date()
        current_date = datetime.now(BEIJING_TZ).strftime('%Y-%m-%d')
        
        print(f"基准日期: {baseline_date}")
        print(f"当前日期: {current_date}")
        
        incremental_data = await prepare_incremental_data(baseline_date)
        orders_data = incremental_data.get('orders', [])
        expense_records = incremental_data.get('expenses', [])
        
        print(f"增量订单数: {len(orders_data)}")
        print(f"增量开销记录数: {len(expense_records)}")
        
        excel_path = await export_incremental_orders_report_to_excel(
            baseline_date,
            current_date,
            orders_data,
            expense_records
        )
        
        if os.path.exists(excel_path):
            file_size = os.path.getsize(excel_path) / 1024
            print(f"✅ Excel已生成: {excel_path}")
            print(f"✅ 文件大小: {file_size:.2f} KB")
            return excel_path
        else:
            print(f"❌ 文件不存在")
            return None
            
    except Exception as e:
        print(f"❌ 测试失败: {e}")
        import traceback
        traceback.print_exc()
        return None


async def test_daily_changes_excel():
    """测试每日数据变更Excel"""
    print("\n" + "=" * 60)
    print("测试2: 每日数据变更Excel")
    print("=" * 60)
    
    try:
        current_date = datetime.now(BEIJING_TZ).strftime('%Y-%m-%d')
        print(f"查询日期: {current_date}")
        
        excel_path = await export_daily_changes_to_excel(current_date)
        
        if os.path.exists(excel_path):
            file_size = os.path.getsize(excel_path) / 1024
            print(f"✅ Excel已生成: {excel_path}")
            print(f"✅ 文件大小: {file_size:.2f} KB")
            return excel_path
        else:
            print(f"❌ 文件不存在")
            return None
            
    except Exception as e:
        print(f"❌ 测试失败: {e}")
        import traceback
        traceback.print_exc()
        return None


async def test_order_table_excel():
    """测试订单总表Excel"""
    print("\n" + "=" * 60)
    print("测试3: 订单总表Excel")
    print("=" * 60)
    
    try:
        # 获取所有有效订单
        valid_orders = await db_operations.get_all_valid_orders()
        
        # 获取当日数据
        date = get_daily_period_date()
        daily_interest = await db_operations.get_daily_interest_total(date)
        completed_orders = await db_operations.get_completed_orders_by_date(date)
        breach_end_orders = await db_operations.get_breach_end_orders_by_date(date)
        daily_summary = await db_operations.get_daily_summary(date)
        
        print(f"有效订单数: {len(valid_orders)}")
        print(f"完成订单数: {len(completed_orders)}")
        print(f"违约完成数: {len(breach_end_orders)}")
        print(f"当日利息: {daily_interest:,.2f}")
        
        excel_path = await export_orders_to_excel(
            valid_orders,
            completed_orders,
            breach_end_orders,
            daily_interest,
            daily_summary
        )
        
        if os.path.exists(excel_path):
            file_size = os.path.getsize(excel_path) / 1024
            print(f"✅ Excel已生成: {excel_path}")
            print(f"✅ 文件大小: {file_size:.2f} KB")
            return excel_path
        else:
            print(f"❌ 文件不存在")
            return None
            
    except Exception as e:
        print(f"❌ 测试失败: {e}")
        import traceback
        traceback.print_exc()
        return None


async def verify_excel_files():
    """验证Excel文件"""
    print("\n" + "=" * 60)
    print("验证Excel文件")
    print("=" * 60)
    
    try:
        from openpyxl import load_workbook
        
        temp_dir = os.path.join(project_root, 'temp')
        if not os.path.exists(temp_dir):
            print(f"⚠️  temp目录不存在")
            return
        
        excel_files = [f for f in os.listdir(temp_dir) if f.endswith('.xlsx')]
        
        if not excel_files:
            print("⚠️  未找到Excel文件")
            return
        
        print(f"找到 {len(excel_files)} 个Excel文件\n")
        
        for file in excel_files[:5]:  # 只检查前5个
            file_path = os.path.join(temp_dir, file)
            try:
                wb = load_workbook(file_path)
                print(f"✅ {file}")
                print(f"   工作表: {', '.join(wb.sheetnames)}")
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    print(f"   - {sheet_name}: {ws.max_row}行 x {ws.max_column}列")
                wb.close()
                print()
            except Exception as e:
                print(f"❌ {file}: {str(e)}\n")
        
    except Exception as e:
        print(f"❌ 验证失败: {e}")


async def check_baseline_ready():
    """检查基准日期是否准备好"""
    print("\n" + "=" * 60)
    print("检查基准日期准备情况")
    print("=" * 60)
    
    try:
        baseline_date = await get_or_create_baseline_date()
        exists = await db_operations.check_baseline_exists()
        
        print(f"基准日期: {baseline_date}")
        print(f"基准日期存在: {exists}")
        
        # 检查是否有数据
        incremental_data = await prepare_incremental_data(baseline_date)
        orders_count = len(incremental_data.get('orders', []))
        expenses_count = len(incremental_data.get('expenses', []))
        
        print(f"\n当前增量数据:")
        print(f"  订单数: {orders_count}")
        print(f"  开销记录数: {expenses_count}")
        
        if orders_count == 0 and expenses_count == 0:
            print(f"\n💡 提示: 当前无增量数据，这是正常的")
            print(f"   如果这是第一次录入，可以:")
            print(f"   1. 设置基准日期为今天")
            print(f"   2. 之后的数据将作为增量数据")
        else:
            print(f"\n✅ 已有增量数据，可以生成报表")
        
        return baseline_date
        
    except Exception as e:
        print(f"❌ 检查失败: {e}")
        import traceback
        traceback.print_exc()
        return None


async def main():
    """主测试函数"""
    print("\n" + "=" * 60)
    print("Excel报表功能完整测试")
    print("=" * 60 + "\n")
    
    try:
        # 初始化数据库
        print("初始化数据库...")
        init_db.init_database()
        print("✅ 数据库初始化完成\n")
        
        # 检查基准日期
        baseline_date = await check_baseline_ready()
        
        # 测试所有Excel报表
        incremental_excel = await test_incremental_report_excel()
        daily_excel = await test_daily_changes_excel()
        order_table_excel = await test_order_table_excel()
        
        # 验证Excel文件
        await verify_excel_files()
        
        # 总结
        print("=" * 60)
        print("测试总结")
        print("=" * 60)
        print(f"✅ 增量订单报表: {'通过' if incremental_excel else '失败'}")
        print(f"✅ 每日数据变更: {'通过' if daily_excel else '失败'}")
        print(f"✅ 订单总表: {'通过' if order_table_excel else '失败'}")
        print(f"✅ 基准日期: {baseline_date if baseline_date else '未设置'}")
        
        print("\n" + "=" * 60)
        print("✅ 所有测试完成！")
        print("=" * 60)
        print("\n💡 下一步:")
        print("1. 检查生成的Excel文件")
        print("2. 验证数据格式和内容")
        print("3. 确认基准日期设置正确")
        print("4. 准备部署到生产环境")
        
    except Exception as e:
        print(f"\n❌ 测试失败: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    asyncio.run(main())

